# build_A.py — Informe A: económico (dinero). Genérico para cualquier semana.
import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from comun_informes import *

d = cargar()
S_ACT = d["semanaActual"]; S_ANT = d["semanaAnterior"]
NA, NP = d["numSemanaActual"], d["numSemanaAnterior"]
ETQ = d["_etiqueta"]
DM = dias_map(d)
prodAct = productos_semana(d, set(S_ACT))
SS = 1.35  # escenario Seguridad Social (+35 % del personal)
HOY = d["hasta"]
tarifa = d.get("tarifaMdna")
ETQ_TARIFA = f"tarifa Mercadona S{tarifa['semana']}" if tarifa else "SIN TARIFA MDNA"

wb = Workbook()

# ═══ Hoja: Por día ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Por día"
titulo(ws, f"Cuenta del día — semanas {NP} (referencia) y {NA}",
       "Ingresos = kg de cada destino × su precio · Coste = naranja (báscula, lotes liquidados) + envase + personal (horas reales) + suministros", 17)
cabecera(ws, 4, ["Fecha", "Kg calibrador", "Producción real (kg)", "Horas línea", "t/h línea", "t/h máquina",
                 "Personas", "Horas trabajadas", "Ingresos €", "Naranja € (conocida)", "Kg sin coste naranja",
                 "Envase €", "Personal €", "Suministros €", "Coste total €", "Neto €", "Neto €/kg"])

def fila_dia(ws, r, dia):
    c1 = poner(ws, r, 1, ETQ[dia["fecha"]])
    if dia["sinParte"] and dia["fecha"] in S_ACT:
        comentario(c1, "Día PROVISIONAL: sin parte diario todavía (kg de los informes del calibrador recibidos hasta ahora). Se completa al regenerar mañana.")
    poner(ws, r, 2, round(dia["kgCalibrador"]), FMT_KG)
    poner(ws, r, 3, round(dia["produccionReal"]), FMT_KG)
    poner(ws, r, 4, dia["horasLinea"], FMT_H)
    poner(ws, r, 5, dia["thEfectiva"], FMT_TH)
    poner(ws, r, 6, dia["thMaquina"], FMT_TH)
    poner(ws, r, 7, dia["presentes"])
    poner(ws, r, 8, dia["horasReales"], FMT_H)
    prods = [p for p in d["productosDia"] if p["fecha"] == dia["fecha"]]
    poner(ws, r, 9, sum(p["ingresoEur"] for p in prods), FMT_EUR)
    poner(ws, r, 10, sum(p["frutaEur"] for p in prods), FMT_EUR)
    poner(ws, r, 11, round(sum(p["kgSinFruta"] for p in prods)), FMT_KG)
    poner(ws, r, 12, sum(p["envaseEur"] for p in prods), FMT_EUR)
    poner(ws, r, 13, dia["costePersonalEur"], FMT_EUR)
    poner(ws, r, 14, dia["suministrosEur"], FMT_EUR)
    poner(ws, r, 15, f"=J{r}+L{r}+M{r}+N{r}", FMT_EUR)
    poner(ws, r, 16, f"=I{r}-O{r}", FMT_EUR)
    poner(ws, r, 17, f"=IFERROR(P{r}/B{r},\"\")", FMT_EURKG)

def fila_total(ws, r, r0, r1, etiqueta):
    poner(ws, r, 1, etiqueta, bold=True, fill=FILL_TOTAL)
    for c in [2, 3, 4, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
        L = get_column_letter(c)
        poner(ws, r, c, f"=SUM({L}{r0}:{L}{r1})", FMT_KG if c in (2, 3, 11) else (FMT_H if c in (4, 8) else FMT_EUR if c >= 9 else None), bold=True, fill=FILL_TOTAL)
    c7 = poner(ws, r, 7, f"=AVERAGE(G{r0}:G{r1})", '0.0', bold=True, fill=FILL_TOTAL)
    comentario(c7, "Personas de MEDIA al día (no la suma de la semana).")
    c5 = poner(ws, r, 5, f"=IFERROR(SUMPRODUCT(E{r0}:E{r1},D{r0}:D{r1})/D{r},\"\")", FMT_TH, bold=True, fill=FILL_TOTAL)
    comentario(c5, "t/h media de la semana ponderada por horas de línea de cada día.")
    poner(ws, r, 6, "", fill=FILL_TOTAL)
    poner(ws, r, 17, f"=IFERROR(P{r}/B{r},\"\")", FMT_EURKG, bold=True, fill=FILL_TOTAL)
    for c in range(1, 18):
        ws.cell(row=r, column=c).border = BORDE_TOTAL

r = 5
poner(ws, r, 1, f"SEMANA {NP} · {rango_texto(S_ANT)} (referencia)", bold=True, fill=FILL_CAB2)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
r += 1
rA0 = r
for f in S_ANT:
    fila_dia(ws, r, DM[f]); r += 1
rA_t = r
fila_total(ws, r, rA0, rA_t - 1, f"TOTAL S{NP}")
r += 2
poner(ws, r, 1, f"SEMANA {NA} · {rango_texto(S_ACT)} (informe)", bold=True, fill=FILL_CAB2)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
r += 1
rB0 = r
for f in S_ACT:
    fila_dia(ws, r, DM[f]); r += 1
rB_t = r
fila_total(ws, r, rB0, rB_t - 1, f"TOTAL S{NA}")
nota(ws, r + 2, "Naranja € = coste all-in de báscula (compra + recolección + transporte + comisión) de los lotes LIQUIDADOS que pasaron ese día. 'Kg sin coste naranja' = lotes aún sin liquidar: el neto real será algo menor.", 17)
nota(ws, r + 3, f"Personal € = horas reales del reloj × coste/hora de cada trabajador (media {d['constantes']['costeHoraMedio']} €/h si no tiene coste cargado). Sin Seguridad Social: ver escenario +35 % en el Resumen.", 17)
anchos(ws, [22, 11, 12, 9, 8, 9, 9, 10, 12, 12, 11, 10, 11, 11, 12, 12, 10])
ws.freeze_panes = "B5"

# ═══ Hoja: Resumen ════════════════════════════════════════════════════════════
rs = wb.create_sheet("Resumen", 0)
titulo(rs, f"Informe económico — Semana {NA} ({rango_texto(S_ACT)})",
       f"Lasarte Cítricos S.L. · generado con datos hasta {ETQ.get(HOY, HOY)} · metodología de Económico → Rentabilidad", 6)
cabecera(rs, 4, ["", "Concepto", f"Semana {NA}", f"Semana {NP} (ref.)", "Δ", ""])
PD = "'Por día'!"
kpis = [
    ("Kg calibrados", f"={PD}B{rB_t}", f"={PD}B{rA_t}", FMT_KG),
    ("Producción real (kg)", f"={PD}C{rB_t}", f"={PD}C{rA_t}", FMT_KG),
    ("Velocidad media línea (t/h)", f"={PD}E{rB_t}", f"={PD}E{rA_t}", FMT_TH),
    ("Ingresos €", f"={PD}I{rB_t}", f"={PD}I{rA_t}", FMT_EUR),
    ("Coste naranja € (lotes liquidados)", f"={PD}J{rB_t}", f"={PD}J{rA_t}", FMT_EUR),
    ("Kg sin coste de naranja (sin liquidar)", f"={PD}K{rB_t}", f"={PD}K{rA_t}", FMT_KG),
    ("Envase €", f"={PD}L{rB_t}", f"={PD}L{rA_t}", FMT_EUR),
    ("Personal € (horas reales)", f"={PD}M{rB_t}", f"={PD}M{rA_t}", FMT_EUR),
    ("Suministros € (estimación)", f"={PD}N{rB_t}", f"={PD}N{rA_t}", FMT_EUR),
    ("Coste total € (sin SS)", f"={PD}O{rB_t}", f"={PD}O{rA_t}", FMT_EUR),
    ("NETO € (sin SS)", f"={PD}P{rB_t}", f"={PD}P{rA_t}", FMT_EUR),
]
r = 5
filas_kpi = {}
for nombre, cA, cB, fmt in kpis:
    poner(rs, r, 2, nombre, bold=(("NETO" in nombre) or ("Coste total" in nombre)))
    poner(rs, r, 3, cA, fmt, bold="NETO" in nombre)
    poner(rs, r, 4, cB, fmt)
    poner(rs, r, 5, f"=IFERROR(C{r}/D{r}-1,\"\")", FMT_PCT)
    filas_kpi[nombre] = r
    r += 1
rP = filas_kpi["Personal € (horas reales)"]
rC = filas_kpi["Coste total € (sin SS)"]
rN = filas_kpi["NETO € (sin SS)"]
poner(rs, r, 2, "Personal con Seguridad Social (+35 %) €")
poner(rs, r, 3, f"=C{rP}*{SS}", FMT_EUR); poner(rs, r, 4, f"=D{rP}*{SS}", FMT_EUR)
comentario(rs.cell(row=r, column=2), "La SS se aproxima como +35 % del coste de nómina (mismo criterio que el informe del 27-jul–4-ago).")
rPss = r; r += 1
poner(rs, r, 2, "NETO € con SS", bold=True)
poner(rs, r, 3, f"=C{rN}-(C{rPss}-C{rP})", FMT_EUR, bold=True)
poner(rs, r, 4, f"=D{rN}-(D{rPss}-D{rP})", FMT_EUR)
poner(rs, r, 5, f"=IFERROR(C{r}/D{r}-1,\"\")", FMT_PCT)
r += 2

poner(rs, r, 2, "Medias €/kg (sobre kg calibrados)", bold=True, fill=FILL_CAB2)
rs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
r += 1
rKG = filas_kpi["Kg calibrados"]
rSin = filas_kpi["Kg sin coste de naranja (sin liquidar)"]
for nombre, fA, fB in [
    ("Precio de venta medio €/kg", f"=C{filas_kpi['Ingresos €']}/C{rKG}", f"=D{filas_kpi['Ingresos €']}/D{rKG}"),
    ("Coste naranja medio €/kg (sobre kg con coste)", f"=C{filas_kpi['Coste naranja € (lotes liquidados)']}/(C{rKG}-C{rSin})", f"=D{filas_kpi['Coste naranja € (lotes liquidados)']}/(D{rKG}-D{rSin})"),
    ("Forfait de confección medio €/kg (envase+personal+suministros)", f"=(C{filas_kpi['Envase €']}+C{rP}+C{filas_kpi['Suministros € (estimación)']})/C{rKG}", f"=(D{filas_kpi['Envase €']}+D{rP}+D{filas_kpi['Suministros € (estimación)']})/D{rKG}"),
    ("Coste total medio €/kg (sin SS)", f"=C{rC}/C{rKG}", f"=D{rC}/D{rKG}"),
    ("Neto medio €/kg (sin SS)", f"=C{rN}/C{rKG}", f"=D{rN}/D{rKG}"),
]:
    poner(rs, r, 2, nombre)
    poner(rs, r, 3, fA, FMT_EURKG)
    poner(rs, r, 4, fB, FMT_EURKG)
    r += 1
r += 1

# Notas calculadas con los datos reales del periodo
kgSinAct = sum(p["kgSinFruta"] for p in d["productosDia"] if p["fecha"] in S_ACT)
kgSinAnt = sum(p["kgSinFruta"] for p in d["productosDia"] if p["fecha"] in S_ANT)
frutaAct = sum(p["frutaEur"] for p in d["productosDia"] if p["fecha"] in S_ACT)
kgConAct = sum(p["kgConFruta"] for p in d["productosDia"] if p["fecha"] in S_ACT)
eurKgFruta = frutaAct / kgConAct if kgConAct else 0.58
real, modelo, kgRealMdna = mdna_real_vs_modelo(d, set(S_ACT))
notas = [
    "El neto está incompleto por los kg sin coste de naranja (lotes sin liquidar en báscula): cuando se liquiden, el neto bajará en su coste.",
    f"OJO al comparar semanas: S{NA} tiene {kgSinAct:,.0f} kg sin liquidar y S{NP} {kgSinAnt:,.0f} — a ~{eurKgFruta:.2f} €/kg son ~{kgSinAct*eurKgFruta:,.0f} € y ~{kgSinAnt*eurKgFruta:,.0f} € pendientes de restar.".replace(",", "."),
]
if tarifa:
    p = tarifa["precios"]
    notas.append(f"Precios Mercadona a {ETQ_TARIFA}: granel {p['mdnaGranel']:.3f} · 3 kg {p['mdna3']:.3f} · 4 kg {p['mdna4']:.3f} · 5 kg {p['mdna5']:.3f} €/kg.")
if real and modelo:
    notas.append(f"Contraste con lo ya facturado por Mercadona esta semana en el ERP ({kgRealMdna:,.0f} kg): {real:.3f} €/kg reales vs {modelo:.3f} del modelo (efecto mezcla de formatos).".replace(",", "."))
notas.append("Otros clientes a precio estándar (empaquetado 0,50 · granel 0,33 · precalibrado 0,33 · industria 0,14 €/kg) salvo ficha con precio propio.")
notas.append("El precalibrado se valora a 0,33 €/kg el día que se aparta y sus kg volverán a contar el día que se confeccionen (doble conteo consciente entre días, igual que la app).")
for a in notas:
    nota(rs, r, "· " + a, 6); r += 1

ch = BarChart(); ch.type = "col"; ch.style = 10
ch.title = f"Semana {NA} — ingresos vs coste por día"
ch.y_axis.title = "€"
datos = Reference(ws, min_col=9, min_row=rB0, max_row=rB_t - 1)
datos2 = Reference(ws, min_col=15, min_row=rB0, max_row=rB_t - 1)
cats = Reference(ws, min_col=1, min_row=rB0, max_row=rB_t - 1)
ch.add_data(datos, titles_from_data=False)
ch.add_data(datos2, titles_from_data=False)
ch.set_categories(cats)
ch.series[0].tx = SeriesLabel(v="Ingresos")
ch.series[1].tx = SeriesLabel(v="Coste total")
ch.series[0].graphicalProperties.solidFill = COLOR["mdna"]
ch.series[1].graphicalProperties.solidFill = COLOR["podrido"]
ch.gapWidth = 60
ch.width = 16; ch.height = 8
rs.add_chart(ch, f"B{r + 1}")  # debajo de las notas: así entra en la página impresa
fin_resumen = r + 20
anchos(rs, [2, 52, 15, 15, 9, 2])

# ═══ Hoja: Por producto ══════════════════════════════════════════════════════
wp = wb.create_sheet("Por producto")
titulo(wp, f"Por producto — semana {NA} acumulada ({rango_texto(S_ACT)})",
       "Kg del calibrador por producto de confección · naranja al coste real de sus lotes · confección repartida por kg ponderado (malla 2,5 · mesa 2,0 · granel 1,0 · industria 0,3)", 21)
cabecera(wp, 4, ["Producto", "Empaque", "Zona", "Destino", "Kg", "% del total", "Precio venta €/kg", "Fuente del precio",
                 "Ingreso €", "Naranja €/kg", "Kg sin coste naranja", "Naranja €", "Envase €/kg", "Envase €",
                 "Confección €", "Confección €/kg", "Forfait confección €/kg", "Coste total €", "Coste €/kg", "NETO €", "NETO €/kg"])
r0 = 5
rT = r0 + len(prodAct)
r = r0
for p in prodAct:
    kgcf = p["kgConFruta"]
    poner(wp, r, 1, p["nombre"])
    poner(wp, r, 2, p["empaque"] or "—")
    poner(wp, r, 3, p["zona"])
    poner(wp, r, 4, DESTINO_ES[p["destino"]])
    poner(wp, r, 5, round(p["kg"]), FMT_KG)
    poner(wp, r, 6, f"=E{r}/E${rT}", FMT_PCT)
    poner(wp, r, 7, p["precioEurKg"] or 0, FMT_EURKG)
    poner(wp, r, 8, p["precioFuente"])
    poner(wp, r, 9, f"=E{r}*G{r}", FMT_EUR)
    cJ = poner(wp, r, 10, (p["frutaEur"] / kgcf) if kgcf > 0 else 0, FMT_EURKG)
    if p["kgSinFruta"] > 0 and kgcf == 0:
        comentario(cJ, "Todo el kg de este producto viene de lotes sin liquidar: coste de naranja aún desconocido (no es 0).")
    poner(wp, r, 11, round(p["kgSinFruta"]), FMT_KG)
    poner(wp, r, 12, f"=(E{r}-K{r})*J{r}", FMT_EUR)
    poner(wp, r, 13, (p["envaseEur"] / p["kg"]) if p["kg"] > 0 else 0, FMT_EURKG)
    poner(wp, r, 14, f"=E{r}*M{r}", FMT_EUR)
    poner(wp, r, 15, p["confeccionEur"], FMT_EUR)
    poner(wp, r, 16, f"=IFERROR(O{r}/E{r},0)", FMT_EURKG)
    poner(wp, r, 17, f"=M{r}+P{r}", FMT_EURKG)
    poner(wp, r, 18, f"=L{r}+N{r}+O{r}", FMT_EUR)
    poner(wp, r, 19, f"=IFERROR(R{r}/E{r},0)", FMT_EURKG)
    poner(wp, r, 20, f"=I{r}-R{r}", FMT_EUR)
    poner(wp, r, 21, f"=IFERROR(T{r}/E{r},0)", FMT_EURKG)
    r += 1
poner(wp, rT, 1, "TOTAL / MEDIA PONDERADA", bold=True, fill=FILL_TOTAL)
for c, fmt in [(5, FMT_KG), (9, FMT_EUR), (11, FMT_KG), (12, FMT_EUR), (14, FMT_EUR), (15, FMT_EUR), (18, FMT_EUR), (20, FMT_EUR)]:
    L = get_column_letter(c)
    poner(wp, rT, c, f"=SUM({L}{r0}:{L}{rT-1})", fmt, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 6, f"=SUM(F{r0}:F{rT-1})", FMT_PCT, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 7, f"=I{rT}/E{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 10, f"=L{rT}/(E{rT}-K{rT})", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 13, f"=N{rT}/E{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 16, f"=O{rT}/E{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 17, f"=M{rT}+P{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 19, f"=R{rT}/E{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
poner(wp, rT, 21, f"=T{rT}/E{rT}", FMT_EURKG, bold=True, fill=FILL_TOTAL)
for c in range(1, 22):
    wp.cell(row=rT, column=c).border = BORDE_TOTAL
nota(wp, rT + 2, "Forfait de confección €/kg = envase + reparto de personal y suministros por kg ponderado. La naranja €/kg es la media real de los lotes de los que salió cada producto ('Kg sin coste' = lotes sin liquidar, no incluidos en el coste).", 21)
nota(wp, rT + 3, "El PRECALIBRADO sale en negativo a propósito: hoy carga toda su naranja y solo vale 0,33 €/kg; el día que se confeccione volverá a contar con su producto final. La INDUSTRIA pierde siempre por la regla de reparto plano de la fruta (decisión del dueño 07-08-2026).", 21)
wp.auto_filter.ref = f"A4:U{rT-1}"
anchos(wp, [40, 26, 11, 24, 10, 9, 10, 22, 11, 10, 11, 11, 9, 10, 11, 10, 11, 12, 9, 12, 10])
wp.freeze_panes = "B5"

# ═══ Hoja: Producto × día ════════════════════════════════════════════════════
wd = wb.create_sheet("Producto × día")
titulo(wd, f"Producto × día (semanas {NP} y {NA})", "Usa el filtro de la fila 4 para ver un día o un producto concreto", 14)
cabecera(wd, 4, ["Fecha", "Semana", "Producto", "Empaque", "Destino", "Kg", "Precio €/kg", "Ingreso €",
                 "Naranja €", "Kg sin coste", "Envase €", "Confección €", "Coste €", "NETO €"])
r = 5
for p in sorted(d["productosDia"], key=lambda x: (x["fecha"], -x["kg"])):
    poner(wd, r, 1, ETQ[p["fecha"]])
    poner(wd, r, 2, f"S{NA}" if p["fecha"] in S_ACT else f"S{NP}")
    poner(wd, r, 3, p["nombre"])
    poner(wd, r, 4, p["empaque"] or "—")
    poner(wd, r, 5, DESTINO_ES[p["destino"]])
    poner(wd, r, 6, round(p["kg"]), FMT_KG)
    poner(wd, r, 7, p["precioEurKg"] or 0, FMT_EURKG)
    poner(wd, r, 8, f"=F{r}*G{r}", FMT_EUR)
    poner(wd, r, 9, p["frutaEur"], FMT_EUR)
    poner(wd, r, 10, round(p["kgSinFruta"]), FMT_KG)
    poner(wd, r, 11, p["envaseEur"], FMT_EUR)
    poner(wd, r, 12, p["confeccionEur"], FMT_EUR)
    poner(wd, r, 13, f"=I{r}+K{r}+L{r}", FMT_EUR)
    poner(wd, r, 14, f"=H{r}-M{r}", FMT_EUR)
    r += 1
wd.auto_filter.ref = f"A4:N{r-1}"
anchos(wd, [10, 8, 40, 26, 24, 10, 10, 11, 11, 10, 10, 11, 11, 12])
wd.freeze_panes = "A5"

# ═══ Hoja: Salidas y ventas ERP ══════════════════════════════════════════════
wv = wb.create_sheet("Salidas y ventas ERP")
titulo(wv, f"Palets dados de alta en el ERP — semana {NA}",
       "Lo que salió de línea cada día con su cliente y su € cuando el albarán está valorado · el día de HOY va hasta la última sincronización del espejo", 9)
cabecera(wv, 4, ["Fecha", "Cliente", "Artículo", "Tipo", "Estado", "Palets", "Kg", "€ facturado", "€/kg"])
agg = {}
for v in d["ventasErp"]:
    if v["fecha"] not in S_ACT:
        continue
    tipo = "interno" if es_interno(v["articulo"]) else "comercial"
    k = (v["fecha"], v["cliente"] or "—", v["articulo"] or "—", tipo, v["estado"])
    a = agg.setdefault(k, [0, 0.0, 0.0, False])
    a[0] += 1; a[1] += v["kg"]
    if v["eur"] is not None:
        a[2] += v["eur"]; a[3] = True
r = 5
r0 = r
for (fecha, cliente, articulo, tipo, estado), (n, kg, eur, con_eur) in sorted(agg.items(), key=lambda x: (x[0][0], -x[1][1])):
    poner(wv, r, 1, ETQ[fecha])
    poner(wv, r, 2, cliente)
    poner(wv, r, 3, articulo)
    poner(wv, r, 4, tipo)
    poner(wv, r, 5, estado)
    poner(wv, r, 6, n)
    poner(wv, r, 7, round(kg), FMT_KG)
    if con_eur and estado == "valorado":
        poner(wv, r, 8, eur, FMT_EUR)
        poner(wv, r, 9, f"=IFERROR(H{r}/G{r},\"\")", FMT_EURKG)
    else:
        poner(wv, r, 8, "—")
        poner(wv, r, 9, "")
    r += 1
rT = r
poner(wv, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
poner(wv, rT, 6, f"=SUM(F{r0}:F{rT-1})", None, bold=True, fill=FILL_TOTAL)
poner(wv, rT, 7, f"=SUM(G{r0}:G{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(wv, rT, 8, f"=SUMIF(E{r0}:E{rT-1},\"valorado\",H{r0}:H{rT-1})", FMT_EUR, bold=True, fill=FILL_TOTAL)
nota(wv, rT + 2, "'Sin albarán' = en cámara/pendiente de venta · 'albarán sin valorar' = enviado y pendiente de precio (consignación / factura posterior) · 'interno' = movimientos a precalibrado o Cítrica, no son venta.", 9)
wv.auto_filter.ref = f"A4:I{rT-1}"
anchos(wv, [10, 30, 34, 11, 17, 8, 10, 12, 9])
wv.freeze_panes = "A5"

# ═══ Hoja: Lotes y naranja ═══════════════════════════════════════════════════
wl = wb.create_sheet("Lotes y naranja")
titulo(wl, f"Lotes que pasaron por línea — semana {NA}", "Coste all-in de báscula por lote · el % podrido e industria son de lo que el calibrador midió esta semana", 10)
cabecera(wl, 4, ["Fecha", "Pasada (lote del calibrador)", "Lote base", "Agricultor", "Kg calibrados", "% podrido", "% industria", "Naranja €/kg", "Coste naranja €", "Estado del coste"])
r = 5
r0 = r
for lt in sorted([x for x in d["lotesDia"] if x["fecha"] in S_ACT], key=lambda x: (x["fecha"], -x["kg"])):
    poner(wl, r, 1, ETQ[lt["fecha"]])
    poner(wl, r, 2, lt["lote"])
    poner(wl, r, 3, lt["lote8"] or "—")
    poner(wl, r, 4, lt["agricultor"] or "—")
    poner(wl, r, 5, round(lt["kg"]), FMT_KG)
    poner(wl, r, 6, lt["pctPodrido"], FMT_PCT)
    poner(wl, r, 7, lt["pctIndustria"], FMT_PCT)
    if lt["frutaEurKg"] is not None:
        poner(wl, r, 8, lt["frutaEurKg"], FMT_EURKG)
        poner(wl, r, 9, f"=E{r}*H{r}", FMT_EUR)
        poner(wl, r, 10, "Liquidado")
    else:
        poner(wl, r, 8, "s/d")
        poner(wl, r, 9, "")
        poner(wl, r, 10, "SIN LIQUIDAR", bold=True)
    r += 1
rT = r
poner(wl, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
poner(wl, rT, 5, f"=SUM(E{r0}:E{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(wl, rT, 9, f"=SUM(I{r0}:I{rT-1})", FMT_EUR, bold=True, fill=FILL_TOTAL)
nota(wl, rT + 2, "Una pasada compuesta («26081302-12 BOX + 26081202-9 BOX») se atribuye al primer lote del código, la misma convención que usa la app.", 10)
anchos(wl, [10, 44, 11, 30, 11, 9, 10, 11, 12, 14])
wl.freeze_panes = "A5"

# ═══ Hoja: Metodología ═══════════════════════════════════════════════════════
wm = wb.create_sheet("Metodología")
titulo(wm, "Metodología, fuentes y avisos", "Todo número de este informe sale de una fuente registrada; lo que falta se dice, no se estima en silencio", 8)
r = 4
def bloque(t):
    global r
    poner(wm, r, 1, t, bold=True, fill=FILL_CAB2, borde=False)
    wm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
def linea(t):
    global r
    nota(wm, r, t, 8)
    r += 1

FUENTE_ES = {"bd": "BD espejo SizerResults", "docx": "DOCX del receptor (validados)", "sin_datos": "SIN DATOS"}
bloque("Fuentes por día (kg del calibrador)")
cabecera(wm, r, ["Día", "Fuente del detalle", "Pasadas", "Kg fuente", "Kg del parte diario", "Diferencia", "", ""], oscura=False); r += 1
for f in S_ANT + S_ACT:
    dia = DM[f]
    poner(wm, r, 1, ETQ[f]); poner(wm, r, 2, FUENTE_ES.get(dia["fuente"], dia["fuente"])); poner(wm, r, 3, dia["nPasadas"])
    poner(wm, r, 4, round(dia["kgCalibrador"]), FMT_KG)
    if dia["sinParte"]:
        poner(wm, r, 5, "sin parte aún")
        poner(wm, r, 6, "")
    else:
        poner(wm, r, 5, round(dia["prodParte"]), FMT_KG)
        poner(wm, r, 6, f"=D{r}-E{r}", FMT_KG)
    r += 1
r += 1
bloque("Precios aplicados (€/kg)")
cabecera(wm, r, ["Destino", "€/kg", "Fuente", "", "", "", "", ""], oscura=False); r += 1
pm = d["preciosMdna"]["aplicados"]
for destino, precio, fuente in [
    ("Mercadona granel 12 kg", pm.get("mdnaGranel"), ETQ_TARIFA + " (última completa facturada)"),
    ("Mercadona malla 3 kg", pm.get("mdna3"), ETQ_TARIFA),
    ("Mercadona girsac 4 kg", pm.get("mdna4"), ETQ_TARIFA),
    ("Mercadona malla 5 kg", pm.get("mdna5"), ETQ_TARIFA),
    ("Empaquetado otros clientes", pm.get("otrosEmp"), "Estándar de la app (proxy ventas junio)"),
    ("Granel otros clientes", pm.get("otrosGranel"), "Estándar de la app"),
    ("Precalibrado", pm.get("prec"), "Estándar de la app (se revaloriza al confeccionarse)"),
    ("Industria", pm.get("industria"), "Dato del dueño (31-07-2026)"),
    ("Podrido / muestras", 0.0, "Sin valor"),
]:
    poner(wm, r, 1, destino); poner(wm, r, 2, precio, FMT_EURKG); poner(wm, r, 3, fuente)
    wm.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1
r += 1
bloque("Constantes de coste")
for t in [
    "Envase €/kg por destino: malla 3 kg 0,0378 · girsac 4 kg 0,045 · malla 5 kg 0,0485 · granel MDNA 0,02 · empaquetado otros 0,04 · granel otros 0,02 (config de mallas de la app).",
    f"Suministros: {d['constantes']['suministrosDia']} €/día (estimación con las facturas de mayo, la misma de la app).",
    f"Personal: horas reales del reloj × coste/hora de la ficha; media {d['constantes']['costeHoraMedio']} €/h para quien no tiene coste cargado. Sin SS; escenario +35 % en el Resumen.",
    "Reparto de personal+suministros entre productos por kg PONDERADO: malla 2,5 · mesa 2,0 · granel 1,0 · industria 0,3 · podrido/prec/muestras no absorben.",
]:
    linea("· " + t)
r += 1
bloque("Avisos de este periodo")
for a in d["avisos"]:
    linea("· " + a)
anchos(wm, [24, 34, 10, 12, 14, 12, 10, 10])

# Maquetación de impresión: el informe se imprime y se lleva en mano.
imprimir_todo(wb, sin_titulos=("Resumen", "Metodología"), areas={"Resumen": f"A1:G{fin_resumen}"})

os.makedirs("salida", exist_ok=True)
NOMBRE = f"salida/Informe economico - semana {NA} ({rango_texto(S_ACT).replace('–','-')}).xlsx"
wb.save(NOMBRE)
print("OK informe A ->", NOMBRE)
