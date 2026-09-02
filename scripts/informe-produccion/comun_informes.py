# comun_informes.py — estilos y utilidades compartidas por los dos informes.
import json
import datetime as dt
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins

RUTA_JSON = "salida/informe-datos.json"

# Paleta validada (dataviz, 6 checks en verde sobre superficie clara)
COLOR = {
    "mdna": "4269D0", "otros": "2F8F4E", "prec": "8A63C9",
    "industria": "B07817", "podrido": "93384F", "gris_ref": "B0B6BE",
}

F_TITULO = Font(name="Arial", size=14, bold=True, color="1A1A2E")
F_SUB = Font(name="Arial", size=9, italic=True, color="666666")
F_CAB = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_CAB2 = Font(name="Arial", size=10, bold=True, color="1A1A2E")
F_NORMAL = Font(name="Arial", size=10)
F_BOLD = Font(name="Arial", size=10, bold=True)
F_NOTA = Font(name="Arial", size=9, color="666666")
FILL_CAB = PatternFill("solid", fgColor="35415C")
FILL_CAB2 = PatternFill("solid", fgColor="E8ECF4")
FILL_TOTAL = PatternFill("solid", fgColor="F2E9D8")
FILL_AVISO = PatternFill("solid", fgColor="FFF4E5")
BORDE_FINO = Border(bottom=Side(style="thin", color="D9D9D9"))
BORDE_TOTAL = Border(top=Side(style="medium", color="35415C"))

FMT_KG = '#,##0'
FMT_EUR = '#,##0" €";[Red]-#,##0" €"'
FMT_EURKG = '0.000;[Red]-0.000'
FMT_PCT = '0.0%'
FMT_TH = '0.0'
FMT_H = '0.0'

DIAS_SEMANA = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
MESES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]

DESTINO_ES = {
    "mdnaGranel": "Mercadona granel 12 kg", "mdna3": "Mercadona malla 3 kg",
    "mdna5": "Mercadona malla 5 kg", "mdna4": "Mercadona girsac 4 kg",
    "otrosEmp": "Empaquetado otros clientes", "otrosGranel": "Granel otros clientes",
    "prec": "Precalibrado (vuelve a línea)", "industria": "Industria",
    "podrido": "Podrido", "muestra": "Muestras",
}


def cargar():
    with open(RUTA_JSON, encoding="utf-8") as f:
        d = json.load(f)
    d["_etiqueta"] = {f: etiqueta_dia(f) for f in d["semanaAnterior"] + d["semanaActual"]}
    return d


def etiqueta_dia(fecha_iso):
    f = dt.date.fromisoformat(fecha_iso)
    return f"{DIAS_SEMANA[f.weekday()]} {f.day:02d}"


def rango_texto(fechas):
    if not fechas:
        return "sin días"
    a, b = dt.date.fromisoformat(fechas[0]), dt.date.fromisoformat(fechas[-1])
    if a == b:
        return f"{a.day} de {MESES[a.month-1]}"
    if a.month == b.month:
        return f"{a.day}–{b.day} {MESES[a.month-1]} {a.year}"
    return f"{a.day} {MESES[a.month-1]} – {b.day} {MESES[b.month-1]} {a.year}"


def titulo(ws, texto, sub, ancho=12):
    ws["A1"] = texto
    ws["A1"].font = F_TITULO
    c = ws["A2"]
    c.value = sub
    c.font = F_SUB
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
    if len(sub) > 130:
        ws.row_dimensions[2].height = 12 * (1 + len(sub) // 130)


def cabecera(ws, fila, cols, oscura=True):
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=fila, column=j, value=c)
        cell.font = F_CAB if oscura else F_CAB2
        cell.fill = FILL_CAB if oscura else FILL_CAB2
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def poner(ws, fila, col, valor, fmt=None, bold=False, fill=None, borde=True):
    cell = ws.cell(row=fila, column=col, value=valor)
    cell.font = F_BOLD if bold else F_NORMAL
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if borde:
        cell.border = BORDE_FINO
    return cell


def nota(ws, fila, texto, ancho=12):
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = F_NOTA
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
    if len(texto) > 120:
        ws.row_dimensions[fila].height = 12 * (1 + len(texto) // 120)


def comentario(cell, texto):
    c = Comment(texto, "Informe de producción")
    c.width = 320
    c.height = 120
    cell.comment = c


def anchos(ws, lista):
    for i, a in enumerate(lista, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a


def productos_semana(d, dias):
    acc = {}
    for p in d["productosDia"]:
        if p["fecha"] not in dias:
            continue
        a = acc.setdefault(p["clave"], {
            "nombre": p["nombre"], "empaque": p["empaque"], "zona": p["zona"],
            "destino": p["destino"], "kg": 0.0, "ingresoEur": 0.0, "frutaEur": 0.0,
            "kgConFruta": 0.0, "kgSinFruta": 0.0, "envaseEur": 0.0, "confeccionEur": 0.0,
            "precioFuente": p["precioFuente"], "precioEurKg": p["precioEurKg"],
        })
        for k in ("kg", "ingresoEur", "frutaEur", "kgConFruta", "kgSinFruta", "envaseEur", "confeccionEur"):
            a[k] += p[k]
        if p["empaque"] and not a["empaque"]:
            a["empaque"] = p["empaque"]
    return sorted(acc.values(), key=lambda x: -x["kg"])


def dias_map(d):
    return {x["fecha"]: x for x in d["dias"]}


def destino_kg_dia(d, dias):
    out = defaultdict(lambda: defaultdict(float))
    ing = defaultdict(float)
    for p in d["productosDia"]:
        if p["fecha"] in dias:
            out[p["destino"]][p["fecha"]] += p["kg"]
            ing[p["destino"]] += p["ingresoEur"]
    return out, ing


def kg_destino(d, fechas):
    out = defaultdict(float)
    for p in d["productosDia"]:
        if p["fecha"] in fechas:
            out[p["destino"]] += p["kg"]
    return out


def mdna_real_vs_modelo(d, dias):
    """€/kg real facturado a Mercadona esta semana (ERP) vs el del modelo (tarifa×mezcla)."""
    kg_r = eur_r = 0.0
    for v in d["ventasErp"]:
        if v["fecha"] in dias and v["estado"] == "valorado" and v["cliente"] and "MERCADONA" in v["cliente"].upper():
            kg_r += v["kg"]; eur_r += v["eur"] or 0
    kg_m = eur_m = 0.0
    for p in d["productosDia"]:
        if p["fecha"] in dias and str(p["destino"]).startswith("mdna"):
            kg_m += p["kg"]; eur_m += p["ingresoEur"]
    real = eur_r / kg_r if kg_r > 0 else None
    modelo = eur_m / kg_m if kg_m > 0 else None
    return real, modelo, kg_r


def es_interno(articulo):
    import re
    t = (articulo or "").upper()
    return bool(re.search(r"\bPRE[C]?\d?\b|PRECAL|CITRICAS?\b|CITRICA\b", t))


def preparar_impresion(ws, titulo_filas="4:4", area=None, una_pagina=False, vertical=False):
    """Deja la hoja lista para imprimir tal cual: A4 apaisado, ajustada al ancho,
    cabecera repetida en cada página y pie con el nombre de la hoja y la página.
    El informe se imprime y se discute en persona: el papel es el formato final.
    Con una_pagina=True la hoja se fuerza a UNA página exacta (paquete de 3)."""
    ws.page_setup.orientation = "portrait" if vertical else "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1 if una_pagina else 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.45, bottom=0.5, header=0.2, footer=0.25)
    ws.print_options.horizontalCentered = True
    if titulo_filas:
        ws.print_title_rows = titulo_filas
    if area:
        ws.print_area = area
    ws.oddFooter.center.text = "&A — pág. &P de &N"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.font = "Arial"


def imprimir_todo(wb, sin_titulos=(), areas=None):
    """Aplica la maquetación de impresión a todas las hojas del libro."""
    areas = areas or {}
    for ws in wb.worksheets:
        preparar_impresion(
            ws,
            titulo_filas=None if ws.title in sin_titulos else "4:4",
            area=areas.get(ws.title),
        )


def cargar_estandar():
    """Estándar de kg/persona acordado (estandar.json editable). None si no existe."""
    import os
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "estandar.json")
    try:
        with open(ruta, encoding="utf-8") as f:
            return json.load(f)
    except OSError:
        return None


def resolver_estandar(est, presentes=None):
    """(suelo, objetivo, regimen) que tocan según los PRESENTES del día.

    Desde el 27-08 el estándar depende del régimen de plantilla (decisión del
    dueño tras el análisis por tipo de día): reducida = media plantilla
    (presentes <= cortePlantillaReducida, el régimen de agosto); completa =
    plantilla entera aunque haya faltas. Sin presentes se asume el régimen
    actual (reducida). Compatible con el estandar.json plano antiguo.
    """
    if est is None:
        return None
    regimenes = est.get("regimenes")
    if not regimenes:
        return (est["kgPersonaSuelo"], est["kgPersonaObjetivo"], None)
    corte = est.get("cortePlantillaReducida", 35)
    nombre = "completa" if (presentes is not None and presentes > corte) else "reducida"
    r = regimenes[nombre]
    return (r["kgPersonaSuelo"], r["kgPersonaObjetivo"], nombre)


def evaluar_estandar(kgp, est, presentes=None):
    """(etiqueta, color hex) del semáforo contra el estándar DEL RÉGIMEN; None si no evaluable."""
    resuelto = resolver_estandar(est, presentes)
    if resuelto is None or kgp is None:
        return None
    suelo, objetivo, regimen = resuelto
    sufijo = ""
    if regimen:
        sufijo = " · listón media plantilla" if regimen == "reducida" else " · listón plantilla completa"
    if kgp >= objetivo:
        return (f"EN OBJETIVO{sufijo}", "2F8F4E")
    if kgp >= suelo:
        return (f"entre suelo y objetivo{sufijo}", "B07817")
    return (f"BAJO EL SUELO — día a explicar{sufijo}", "93384F")
