# build_dia_print.py — parte DIARIO imprimible: 2 páginas exactas, sin euros.
# Pág. 1: el día (rendimiento, zonas con arranque y las personas con nombres).
# Pág. 2: la fruta (lotes/fincas que se echaron y productos que salieron).
# Uso: python build_dia_print.py [--fecha=YYYY-MM-DD]   (por defecto, el día 'hasta' del JSON)
import os
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from comun_informes import *
from openpyxl.styles import Font

d = cargar()
arg = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--fecha=")), None)
F = arg or d["hasta"]
DM = dias_map(d)
if F not in DM:
    print(f"sin datos del {F}: no se genera el parte diario imprimible")
    sys.exit(0)
dia = DM[F]
if (dia["kgCalibrador"] or 0) < 500:
    print(f"{F} sin producción: no se genera el parte diario imprimible")
    sys.exit(0)

S_ANT = d["semanaAnterior"]
NP = d["numSemanaAnterior"]
NA = d["numSemanaActual"]
ETQF = etiqueta_dia(F)
nP_d = max(len(S_ANT), 1)

prods = [p for p in d["productosDia"] if p["fecha"] == F]
kgTot = sum(p["kg"] for p in prods) or 1
kgd = defaultdict(float)
for p in prods:
    kgd[p["destino"]] += p["kg"]

def arranque_dia(f):
    return sum(1 for p in DM[f]["detallePersonas"] if p["tipo"] == "tratamiento")

MAPA_ZONA = {"Mesas": "Envasadoras", "Mallas": "Mallas", "Graneleras": "Graneleras", "Industria": "Industria"}
def kg_grupo(f):
    out = defaultdict(float)
    for p in d["productosDia"]:
        if p["fecha"] == f:
            g = MAPA_ZONA.get(p["zona"])
            if g:
                out[g] += p["kg"]
    return out

# medias de la semana anterior como referencia del día
def media_ant(clave):
    vals = [DM[f][clave] for f in S_ANT if DM[f].get(clave) is not None]
    return sum(vals) / len(vals) if vals else None

wb = Workbook()

# ═══ PÁGINA 1 · El día ═══════════════════════════════════════════════════════
w1 = wb.active
w1.title = "1 · El día"
titulo(w1, f"Parte de rendimiento — {ETQF} ({F}) · semana {NA}",
       "Lasarte Cítricos S.L. · el día tal cual fue · producción real = kg calibrados − fruta de mujeres − reciclado · sin euros", 9)

cabecera(w1, 4, ["El día en números", "Valor", "", "", "", "", "", "", ""])
r = 5
arrF = arranque_dia(F)
prAnt = media_ant("produccionReal"); compAnt = media_ant("computables")
filas1 = [
    ("Kg calibrados", dia["kgCalibrador"], media_ant("kgCalibrador"), FMT_KG),
    ("Producción real (kg)", dia["produccionReal"], prAnt, FMT_KG),
    ("Fruta de mujeres (kg)", dia["mujeres"], media_ant("mujeres"), FMT_KG),
    ("Reciclado de mallas (kg)", dia["z1"] + dia["z2"], (media_ant("z1") or 0) + (media_ant("z2") or 0), FMT_KG),
    ("Personas presentes", dia["presentes"], media_ant("presentes"), '0.0'),
    ("· computan kg/persona", dia["computables"], compAnt, '0.0'),
    ("kg / persona", dia["kgPersona"], (prAnt / compAnt) if prAnt and compAnt else None, FMT_KG),
    ("kg / hora·persona", (dia["produccionReal"] / dia["horasReales"]) if dia["horasReales"] else None,
     (sum(DM[f]["produccionReal"] for f in S_ANT) / sum(DM[f]["horasReales"] for f in S_ANT)) if sum(DM[f]["horasReales"] for f in S_ANT) else None, FMT_KG),
    ("Velocidad línea (t/h)", dia["thEfectiva"], media_ant("thEfectiva"), FMT_TH),
    ("Palets dados de alta (kg)", dia["paletsBrutos"], media_ant("paletsBrutos"), FMT_KG),
    ("% podrido + muestras", (kgd.get("podrido", 0) + kgd.get("muestra", 0)) / kgTot, None, FMT_PCT),
    ("% industria", kgd.get("industria", 0) / kgTot, None, FMT_PCT),
    ("% precalibrado", kgd.get("prec", 0) / kgTot, None, FMT_PCT),
    ("% a Mercadona", sum(kgd.get(k, 0) for k in ("mdna3", "mdna4", "mdna5", "mdnaGranel")) / kgTot, None, FMT_PCT),
]
EST = cargar_estandar()
for nombre, v, ref, fmt in filas1:
    poner(w1, r, 1, nombre, bold=nombre in ("kg / persona", "Producción real (kg)"))
    poner(w1, r, 2, round(v, 1) if isinstance(v, (int, float)) and fmt in ('0.0',) else v, fmt, bold=nombre == "kg / persona")
    r += 1
    if nombre == "kg / persona" and EST:
        semaforo = evaluar_estandar(v, EST, dia.get("presentes"))
        if semaforo:
            etiqueta_sem, color_sem = semaforo
            suelo_est, objetivo_est, _regimen = resolver_estandar(EST, dia.get("presentes"))
            poner(w1, r, 1, "   contra el estándar")
            texto_sem = f"{etiqueta_sem}   (suelo {suelo_est:,} · objetivo {objetivo_est:,})".replace(",", ".")
            c = poner(w1, r, 2, texto_sem, bold=True)
            c.font = Font(name="Arial", size=10, bold=True, color=color_sem)
            w1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            r += 1
r += 1

poner(w1, r, 1, "ZONAS DEL DÍA (cada zona = sus personas + el arranque)", bold=True, fill=FILL_CAB2)
w1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
cabecera(w1, r, ["Zona", "Kg", "Personas de la zona", "+ Arranque", "Total", "kg / persona", "", "", ""], oscura=False)
r += 1
kgz = kg_grupo(F)
arrAntM = sum(arranque_dia(f) for f in S_ANT) / nP_d if S_ANT else 0
for g in ["Mallas", "Envasadoras", "Graneleras", "Industria"]:
    pz = 0 if g == "Industria" else dia["personasPorGrupo"].get(g, 0)
    kgzAnt = sum(kg_grupo(f).get(g, 0) for f in S_ANT) / nP_d if S_ANT else 0
    pzAntM = (sum(DM[f]["personasPorGrupo"].get(g, 0) for f in S_ANT) / nP_d) if (S_ANT and g != "Industria") else 0
    kgpAnt = kgzAnt / (pzAntM + arrAntM) if (pzAntM + arrAntM) > 0 else None
    if kgz.get(g, 0) <= 0 and pz <= 0:
        continue
    poner(w1, r, 1, g if g != "Industria" else "Industria (la hace el arranque)", bold=True)
    poner(w1, r, 2, round(kgz.get(g, 0)), FMT_KG)
    poner(w1, r, 3, pz)
    poner(w1, r, 4, arrF)
    poner(w1, r, 5, f"=C{r}+D{r}")
    poner(w1, r, 6, f"=IFERROR(B{r}/E{r},\"—\")", FMT_KG, bold=True)
    r += 1
r += 1

poner(w1, r, 1, f"LAS PERSONAS DE AYER — {dia['presentes']} presentes, {round(dia['horasReales'])} horas (días de esta semana entre paréntesis)", bold=True, fill=FILL_CAB2)
w1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
grupos_dia = {}
dias_semana = {}
for f in d["semanaActual"]:
    for p in DM[f]["detallePersonas"]:
        dias_semana.setdefault(p["nombre"], 0)
        dias_semana[p["nombre"]] += 1
for p in sorted(dia["detallePersonas"], key=lambda x: x["nombre"]):
    clave = p["grupo"] or ("Arranque (línea)" if p["tipo"] == "tratamiento" else ("No computan" if p["tipo"] == "no_computa" else "Otros"))
    grupos_dia.setdefault(clave, []).append(f"{p['nombre']} ({dias_semana.get(p['nombre'], 1)})")
for grupo in ["Arranque (línea)", "Mallas", "Envasadoras", "Graneleras", "Industria", "No computan", "Otros"]:
    lista = grupos_dia.get(grupo)
    if not lista:
        continue
    poner(w1, r, 1, f"{grupo} — {len(lista)}", bold=True)
    texto = "; ".join(lista)
    c = poner(w1, r, 2, texto)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    w1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    w1.row_dimensions[r].height = 13 * (1 + len(texto) // 105)
    r += 1
avisos = [a for a in d["avisos"] if (F in a or not a[:4].isdigit()) and not any(x in a for x in ("€", "facturar", "tarifa", "coste"))]
if dia["sinParte"]:
    avisos.insert(0, "Sin parte diario todavía: producción real provisional (sin reciclados ni podrido de bolsa descontados).")
if avisos:
    r += 1
    nota(w1, r, "REVISAR: " + " · ".join(avisos), 9)
anchos(w1, [30, 12, 13, 11, 9, 12, 13, 9, 9])

# ═══ PÁGINA 2 · La fruta del día ═════════════════════════════════════════════
w2 = wb.create_sheet("2 · La fruta")
titulo(w2, f"La fruta de {ETQF} — lo que se echó y lo que salió",
       "Lotes con su productor y finca (báscula) · productos con su empaque · el % podrido e industria es lo que midió el calibrador en cada lote", 9)

cabecera(w2, 4, ["Lote echado (pasada)", "Lote base", "Productor", "Finca", "Variedad (báscula)", "Entrada", "Kg", "% podrido", "% industria"])
r = 5; r0 = r
lotes = sorted([x for x in d["lotesDia"] if x["fecha"] == F], key=lambda x: -x["kg"])
for lt in lotes:
    poner(w2, r, 1, lt["lote"])
    poner(w2, r, 2, lt["lote8"] or "—")
    poner(w2, r, 3, lt["agricultor"] or "—")
    poner(w2, r, 4, lt.get("finca") or "—")
    poner(w2, r, 5, lt.get("articulo") or "—")
    fe = lt.get("fechaEntrada")
    poner(w2, r, 6, f"{fe[8:10]}-{fe[5:7]}" if fe else "—")
    poner(w2, r, 7, round(lt["kg"]), FMT_KG)
    poner(w2, r, 8, lt["pctPodrido"], FMT_PCT)
    poner(w2, r, 9, lt["pctIndustria"], FMT_PCT)
    r += 1
poner(w2, r, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
poner(w2, r, 7, f"=SUM(G{r0}:G{r-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
rT = r
r += 2

poner(w2, r, 1, "PRODUCTOS QUE SALIERON", bold=True, fill=FILL_CAB2)
w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
cabecera(w2, r, ["Producto", "Empaque", "Destino", "Kg", "% del día", "", "Categorías del calibrador", "Kg", "%"], oscura=False)
r += 1
rP0 = r
lista_prods = sorted(prods, key=lambda x: -x["kg"])
cats = sorted([c for c in d["categoriasDia"] if c["fecha"] == F], key=lambda x: -x["kg"])
nfilas = max(len(lista_prods), len(cats))
for i in range(nfilas):
    if i < len(lista_prods):
        p = lista_prods[i]
        poner(w2, r, 1, p["nombre"])
        poner(w2, r, 2, (p["empaque"] or "—")[:26])
        poner(w2, r, 3, DESTINO_ES[p["destino"]])
        poner(w2, r, 4, round(p["kg"]), FMT_KG)
        poner(w2, r, 5, p["kg"] / kgTot, FMT_PCT)
    if i < len(cats):
        poner(w2, r, 7, cats[i]["clase"])
        poner(w2, r, 8, round(cats[i]["kg"]), FMT_KG)
        poner(w2, r, 9, cats[i]["kg"] / kgTot, FMT_PCT)
    r += 1
poner(w2, r, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
poner(w2, r, 4, f"=SUM(D{rP0}:D{r-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(w2, r, 8, f"=SUM(H{rP0}:H{r-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
r += 2
nota(w2, r, "Una pasada compuesta («26081302-12 BOX + 26081202-9 BOX») se atribuye al primer lote del código, la misma convención que la app. La variedad y la finca son las de la entrada de báscula del lote.", 9)
anchos(w2, [32, 16, 24, 18, 21, 7, 13, 9, 9])

for hoja in wb.worksheets:
    preparar_impresion(hoja, titulo_filas=None, una_pagina=True)

os.makedirs("salida", exist_ok=True)
NOMBRE = f"salida/IMPRIMIR - dia {F}.xlsx"
wb.save(NOMBRE)
print("OK parte diario imprimible ->", NOMBRE)
