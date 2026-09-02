# build_B_print.py — paquete de IMPRESIÓN del rendimiento: 3 páginas exactas.
# Pág. 1: la semana en números (resumen + por día). Pág. 2: zonas y personas
# con nombres. Pág. 3: mezcla de fruta + las reglas en corto. Sin euros.
# El Excel completo (build_B) queda para pantalla; ESTO es lo que se imprime.
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from comun_informes import *
from openpyxl.styles import Font
EST = cargar_estandar()

d = cargar()
S_ACT = d["semanaActual"]; S_ANT = d["semanaAnterior"]
NA, NP = d["numSemanaActual"], d["numSemanaAnterior"]
ETQ = d["_etiqueta"]
TODOS = S_ANT + S_ACT
DM = dias_map(d)

def prods(fechas):
    return [p for p in d["productosDia"] if p["fecha"] in fechas]

def kg_dest(fechas):
    return kg_destino(d, fechas)

def kg_grupo_dia(fecha):
    mapa = {"Mesas": "Envasadoras", "Mallas": "Mallas", "Graneleras": "Graneleras", "Industria": "Industria"}
    out = defaultdict(float)
    for p in prods([fecha]):
        g = mapa.get(p["zona"])
        if g:
            out[g] += p["kg"]
    return out

def arranque_dia(f):
    return sum(1 for p in DM[f]["detallePersonas"] if p["tipo"] == "tratamiento")

GRUPOS = ["Mallas", "Envasadoras", "Graneleras", "Industria"]
nA_d = max(len(S_ACT), 1); nP_d = max(len(S_ANT), 1)

def pct_dest(fechas, claves):
    kgd = kg_dest(fechas)
    tot = sum(kgd.values())
    return sum(kgd.get(k, 0) for k in claves) / tot if tot else None

wb = Workbook()

# ═══ PÁGINA 1 · La semana en números ═════════════════════════════════════════
w1 = wb.active
w1.title = "1 · Semana"
titulo(w1, f"Rendimiento — Semana {NA} ({rango_texto(S_ACT)})",
       f"Lasarte Cítricos S.L. · comparada con la semana {NP} ({rango_texto(S_ANT)}) · producción real = kg calibrados − fruta de mujeres − reciclado", 12)

# Por día (lo esencial)
cabecera(w1, 4, ["Día", "Kg calibrados", "Producción real", "Personas", "Computan", "kg/persona", "kg/h·persona", "t/h línea", "% podrido", "% industria", "% precalibr.", "% a palets"])
r = 5; r0 = r
for f in S_ACT:
    dia = DM[f]
    kgd = kg_dest([f]); tot = sum(kgd.values()) or 1
    poner(w1, r, 1, ETQ[f])
    poner(w1, r, 2, round(dia["kgCalibrador"]), FMT_KG)
    poner(w1, r, 3, round(dia["produccionReal"]), FMT_KG)
    poner(w1, r, 4, dia["presentes"])
    poner(w1, r, 5, dia["computables"])
    celda_kgp = poner(w1, r, 6, f"=IFERROR(C{r}/E{r},\"\")", FMT_KG)
    semaforo = evaluar_estandar((dia["produccionReal"] / dia["computables"]) if dia["computables"] else None, EST, dia.get("presentes"))
    if semaforo:
        celda_kgp.font = Font(name="Arial", size=10, bold=True, color=semaforo[1])
    poner(w1, r, 7, dia["horasReales"] and round(dia["produccionReal"] / dia["horasReales"]), FMT_KG)
    poner(w1, r, 8, dia["thEfectiva"], FMT_TH)
    poner(w1, r, 9, (kgd.get("podrido", 0) + kgd.get("muestra", 0)) / tot, FMT_PCT)
    poner(w1, r, 10, kgd.get("industria", 0) / tot, FMT_PCT)
    poner(w1, r, 11, kgd.get("prec", 0) / tot, FMT_PCT)
    poner(w1, r, 12, (dia["paletsBrutos"] / dia["kgCalibrador"]) if dia["paletsBrutos"] else None, FMT_PCT)
    r += 1
rT = r
poner(w1, rT, 1, f"TOTAL S{NA}", bold=True, fill=FILL_TOTAL)
poner(w1, rT, 2, f"=SUM(B{r0}:B{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(w1, rT, 3, f"=SUM(C{r0}:C{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(w1, rT, 4, f"=AVERAGE(D{r0}:D{rT-1})", '0.0', bold=True, fill=FILL_TOTAL)
poner(w1, rT, 5, f"=AVERAGE(E{r0}:E{rT-1})", '0.0', bold=True, fill=FILL_TOTAL)
poner(w1, rT, 6, f"=IFERROR(C{rT}/(E{rT}*{len(S_ACT)}),\"\")", FMT_KG, bold=True, fill=FILL_TOTAL)
horasTot = sum(DM[f]["horasReales"] for f in S_ACT)
poner(w1, rT, 7, round(sum(DM[f]["produccionReal"] for f in S_ACT) / horasTot) if horasTot else None, FMT_KG, bold=True, fill=FILL_TOTAL)
kgH = sum(DM[f]["kgCalibrador"] for f in S_ACT if DM[f]["horasLinea"]); hL = sum(DM[f]["horasLinea"] or 0 for f in S_ACT)
poner(w1, rT, 8, (kgH / 1000 / hL) if hL else None, FMT_TH, bold=True, fill=FILL_TOTAL)
poner(w1, rT, 9, pct_dest(S_ACT, ["podrido", "muestra"]), FMT_PCT, bold=True, fill=FILL_TOTAL)
poner(w1, rT, 10, pct_dest(S_ACT, ["industria"]), FMT_PCT, bold=True, fill=FILL_TOTAL)
poner(w1, rT, 11, pct_dest(S_ACT, ["prec"]), FMT_PCT, bold=True, fill=FILL_TOTAL)
paletsTot = sum(DM[f]["paletsBrutos"] or 0 for f in S_ACT)
poner(w1, rT, 12, paletsTot / sum(DM[f]["kgCalibrador"] for f in S_ACT), FMT_PCT, bold=True, fill=FILL_TOTAL)
for c in range(1, 13):
    w1.cell(row=rT, column=c).border = BORDE_TOTAL

# Resumen semana vs semana
r = rT + 2
poner(w1, r, 1, f"LA SEMANA {NA} FRENTE A LA {NP}", bold=True, fill=FILL_CAB2)
w1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
r += 1
cabecera(w1, r, ["", f"Semana {NA}", f"Semana {NP}", "Δ", "", "", f"Semana {NA}", f"Semana {NP}", "Δ", "", "", ""], oscura=False)
r += 1
def resumen_semana(fechas, n):
    kg = sum(DM[f]["kgCalibrador"] for f in fechas)
    pr = sum(DM[f]["produccionReal"] for f in fechas)
    comp = sum(DM[f]["computables"] for f in fechas)
    horas = sum(DM[f]["horasReales"] for f in fechas)
    return {"kg": kg, "pr": pr, "kgp": pr / comp if comp else None, "kgh": pr / horas if horas else None,
            "pers": sum(DM[f]["presentes"] for f in fechas) / n, "comp": comp / n}
A = resumen_semana(S_ACT, nA_d); P = resumen_semana(S_ANT, nP_d)
izq = [("Kg calibrados", A["kg"], P["kg"], FMT_KG), ("Producción real", A["pr"], P["pr"], FMT_KG),
       ("kg/persona (día normal)", A["kgp"], P["kgp"], FMT_KG), ("kg/hora·persona", A["kgh"], P["kgh"], FMT_KG)]
der = [("Personas medias/día", A["pers"], P["pers"], '0.0'), ("· computan kg/p", A["comp"], P["comp"], '0.0'),
       ("% a Mercadona", pct_dest(S_ACT, ["mdna3", "mdna4", "mdna5", "mdnaGranel"]), pct_dest(S_ANT, ["mdna3", "mdna4", "mdna5", "mdnaGranel"]), FMT_PCT),
       ("% otros clientes", pct_dest(S_ACT, ["otrosEmp", "otrosGranel"]), pct_dest(S_ANT, ["otrosEmp", "otrosGranel"]), FMT_PCT)]
for i in range(4):
    n1, vA1, vP1, f1 = izq[i]
    poner(w1, r, 1, n1); poner(w1, r, 2, vA1, f1); poner(w1, r, 3, vP1, f1)
    poner(w1, r, 4, f"=IFERROR(B{r}/C{r}-1,\"\")", FMT_PCT)
    n2, vA2, vP2, f2 = der[i]
    c6 = poner(w1, r, 6, n2)
    w1.merge_cells(start_row=r, start_column=6, end_row=r, end_column=6)
    poner(w1, r, 7, vA2, f2); poner(w1, r, 8, vP2, f2)
    poner(w1, r, 9, f"=IFERROR(G{r}/H{r}-1,\"\")", FMT_PCT)
    r += 1
r += 1
if EST:
    dias_bajo = sum(1 for f in S_ACT if DM[f]["computables"] and (DM[f]["produccionReal"] / DM[f]["computables"]) < EST["kgPersonaSuelo"])
    nota(w1, r, f"ESTÁNDAR acordado: suelo {EST['kgPersonaSuelo']:,} · objetivo {EST['kgPersonaObjetivo']:,} kg/persona (colores en la columna kg/persona). Esta semana: {dias_bajo} día(s) bajo el suelo.".replace(",", "."), 12)
    r += 1
nota(w1, r, "Antes de leer el kg/persona como productividad, mirar % podrido / industria / precalibrado: la fruta que viene del campo manda tanto como la gente. Un día se compara bien con otro de mezcla parecida.", 12)
avisos = [a for a in d["avisos"] if not any(x in a for x in ("€", "facturar", "tarifa", "coste"))]
if avisos:
    r += 1
    nota(w1, r, "REVISAR: " + " · ".join(avisos), 12)
anchos(w1, [20, 11, 12, 10, 10, 11, 12, 8, 9, 10, 11, 9])
w1.column_dimensions["F"].width = 20  # etiquetas del bloque derecho, completas

# ═══ PÁGINA 2 · Zonas y personas ═════════════════════════════════════════════
w2 = wb.create_sheet("2 · Zonas y personas")
titulo(w2, f"Zonas y personas — semana {NA} (cada zona = sus personas + el arranque)",
       "El arranque (encargadas, producción, aéreo, tría, volcador, carretillas, transpaletas, mantenimiento) corre la línea: cuenta en todas las zonas, y la industria la produce él solo.", 9)
cabecera(w2, 4, ["Zona · un día normal", "Kg al día", "Personas de la zona", "+ Arranque", "Total", "kg / persona", f"S{NP}: kg/pers.", "Δ", "Kg semana"])
arrAm = sum(arranque_dia(f) for f in S_ACT) / nA_d
arrPm = sum(arranque_dia(f) for f in S_ANT) / nP_d if S_ANT else 0
r = 5
for g in GRUPOS:
    kgA = sum(kg_grupo_dia(f).get(g, 0) for f in S_ACT)
    pzAm = sum(DM[f]["personasPorGrupo"].get(g, 0) for f in S_ACT) / nA_d
    kgP = sum(kg_grupo_dia(f).get(g, 0) for f in S_ANT)
    pzPm = sum(DM[f]["personasPorGrupo"].get(g, 0) for f in S_ANT) / nP_d if S_ANT else 0
    kgpP = (kgP / nP_d) / (pzPm + arrPm) if (pzPm + arrPm) > 0 else None
    poner(w2, r, 1, g if g != "Industria" else "Industria (la hace el arranque)", bold=True)
    poner(w2, r, 2, round(kgA / nA_d), FMT_KG)
    poner(w2, r, 3, round(pzAm, 1), '0.0')
    poner(w2, r, 4, round(arrAm, 1), '0.0')
    poner(w2, r, 5, f"=C{r}+D{r}", '0.0')
    poner(w2, r, 6, f"=IFERROR(B{r}/E{r},\"—\")", FMT_KG, bold=True)
    poner(w2, r, 7, kgpP, FMT_KG)
    poner(w2, r, 8, f"=IFERROR(F{r}/G{r}-1,\"\")", FMT_PCT)
    poner(w2, r, 9, round(kgA), FMT_KG)
    r += 1
r += 1
nota(w2, r, "Personas en media al día (un 9,6 = un día faltó una). Las zonas no se suman entre sí: el arranque cuenta en todas a propósito — responde a '¿cuánta gente hace falta de verdad para producir este kg?'.", 9)
r += 2

# Personas por día por grupo
poner(w2, r, 1, "PERSONAS POR DÍA", bold=True, fill=FILL_CAB2)
w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
cabecera(w2, r, ["Día", "Arranque", "Mallas", "Envasad.", "Granel.", "No computan", "Total computan", "Horas totales", ""], oscura=False)
r += 1
for f in S_ACT:
    dia = DM[f]
    poner(w2, r, 1, ETQ[f])
    poner(w2, r, 2, arranque_dia(f))
    for j, g in enumerate(["Mallas", "Envasadoras", "Graneleras"]):
        poner(w2, r, 3 + j, dia["personasPorGrupo"].get(g, 0))
    poner(w2, r, 6, dia["presentes"] - dia["computables"])
    poner(w2, r, 7, dia["computables"], None, bold=True)
    poner(w2, r, 8, dia["horasReales"], FMT_H)
    r += 1
r += 1

# Quién es quién
poner(w2, r, 1, "QUIÉN ES QUIÉN (días trabajados entre paréntesis)", bold=True, fill=FILL_CAB2)
w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
quien = {}
for f in S_ACT:
    for p in DM[f]["detallePersonas"]:
        clave = p["grupo"] or ("Arranque (línea)" if p["tipo"] == "tratamiento" else ("No computan" if p["tipo"] == "no_computa" else "Otros"))
        quien.setdefault(clave, {}).setdefault(p["nombre"], 0)
        quien[clave][p["nombre"]] += 1
for grupo in ["Arranque (línea)", "Mallas", "Envasadoras", "Graneleras", "Industria", "No computan", "Otros"]:
    personas = quien.get(grupo)
    if not personas:
        continue
    poner(w2, r, 1, f"{grupo} — {len(personas)}", bold=True)
    lista = "; ".join(f"{n} ({v})" for n, v in sorted(personas.items()))
    c = poner(w2, r, 2, lista)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    w2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    w2.row_dimensions[r].height = 13 * (1 + len(lista) // 105)
    r += 1
anchos(w2, [28, 11, 13, 11, 10, 11, 12, 11, 11])

# ═══ PÁGINA 3 · La fruta y las reglas ════════════════════════════════════════
w3 = wb.create_sheet("3 · Fruta y reglas")
titulo(w3, f"Dónde fue la fruta — semana {NA} (y las reglas del informe)",
       "Kg del calibrador por destino y por categoría · la mezcla explica el rendimiento tanto como las personas", 10)
cabecera(w3, 4, ["Destino"] + [ETQ[f] for f in S_ACT] + [f"Total S{NA}", f"% S{NA}", f"% S{NP}", "Δ p.p."])
destA, _ = destino_kg_dia(d, set(S_ACT))
kgTotA = sum(sum(v.values()) for v in destA.values()) or 1
kgDestP = kg_dest(S_ANT); kgTotP = sum(kgDestP.values()) or 1
nD = len(S_ACT)
r = 5
for dst in d["constantes"]["destinosOrden"]:
    kgT = sum(destA.get(dst, {}).values())
    if kgT <= 0 and kgDestP.get(dst, 0) <= 0:
        continue
    poner(w3, r, 1, DESTINO_ES[dst])
    for j, f in enumerate(S_ACT):
        poner(w3, r, 2 + j, round(destA.get(dst, {}).get(f, 0)), FMT_KG)
    poner(w3, r, 2 + nD, round(kgT), FMT_KG)
    poner(w3, r, 3 + nD, kgT / kgTotA, FMT_PCT)
    poner(w3, r, 4 + nD, kgDestP.get(dst, 0) / kgTotP, FMT_PCT)
    poner(w3, r, 5 + nD, kgT / kgTotA - kgDestP.get(dst, 0) / kgTotP, '0.0%;[Red]-0.0%')
    r += 1
r += 1

poner(w3, r, 1, "CATEGORÍAS DEL CALIBRADOR (kg de la semana)", bold=True, fill=FILL_CAB2)
w3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
r += 1
cat = defaultdict(lambda: defaultdict(float))
for c in d["categoriasDia"]:
    cat[c["clase"]][c["fecha"]] += c["kg"]
catA = {k: sum(v for f, v in vv.items() if f in S_ACT) for k, vv in cat.items()}
catP = {k: sum(v for f, v in vv.items() if f in S_ANT) for k, vv in cat.items()}
totA = sum(catA.values()) or 1; totP = sum(catP.values()) or 1
clases = sorted(catA, key=lambda k: -catA[k])[:11]
cabecera(w3, r, ["Clase", f"Kg S{NA}", f"% S{NA}", f"% S{NP}", "Δ p.p.", "", "", "", "", ""], oscura=False)
r += 1
for cl in clases:
    poner(w3, r, 1, cl)
    poner(w3, r, 2, round(catA[cl]), FMT_KG)
    poner(w3, r, 3, catA[cl] / totA, FMT_PCT)
    poner(w3, r, 4, catP.get(cl, 0) / totP, FMT_PCT)
    poner(w3, r, 5, catA[cl] / totA - catP.get(cl, 0) / totP, '0.0%;[Red]-0.0%')
    r += 1
r += 1
poner(w3, r, 1, "LAS REGLAS, EN CORTO", bold=True, fill=FILL_CAB2)
w3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
r += 1
for t in [
    "· Producción real = kg del calibrador − fruta de mujeres − reciclado de mallas. kg/persona = producción real ÷ presentes que computan (fuera oficina, carga/descarga y turno de tarde).",
    "· Cada zona cuenta sus personas + las del arranque; la industria la produce el propio arranque. Por eso las zonas no se suman entre sí.",
    "· kg/hora·persona quita el efecto de jornadas desiguales. La t/h de línea mide la máquina, no a la gente.",
    "· Lo que NO depende de la gente: el podrido y el verde vienen del campo; el precalibrado contará el día que se confeccione; las averías bajan la t/h.",
    "· Personas y zonas salen de la ficha de cada trabajador en la app: si alguien está mal clasificado, se corrige su ficha y el siguiente informe sale bien.",
    "· Datos: calibrador (BD o sus informes DOCX validados), reloj de presencia y ERP (solo lectura). Lo que falta se avisa, nunca se rellena en silencio.",
]:
    nota(w3, r, t, 10)
    r += 1
anchos(w3, [26, 10, 10, 10, 10, 10, 11, 8, 8, 8])

# ═══ Impresión: 3 páginas exactas ════════════════════════════════════════════
for hoja in wb.worksheets:
    preparar_impresion(hoja, titulo_filas=None, una_pagina=True)

os.makedirs("salida", exist_ok=True)
NOMBRE = f"salida/IMPRIMIR - rendimiento semana {NA} ({rango_texto(S_ACT).replace(chr(8211), chr(45))}).xlsx"
wb.save(NOMBRE)
print("OK paquete de impresion ->", NOMBRE)
