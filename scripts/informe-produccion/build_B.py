# build_B.py — Informe B: rendimiento (personas, zonas, categorías, destinos). Genérico.
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break
from comun_informes import *

d = cargar()
S_ACT = d["semanaActual"]; S_ANT = d["semanaAnterior"]
NA, NP = d["numSemanaActual"], d["numSemanaAnterior"]
ETQ = d["_etiqueta"]
TODOS = S_ANT + S_ACT
DM = dias_map(d)

def prods(fechas):
    return [p for p in d["productosDia"] if p["fecha"] in fechas]

def kg_dest(fechas):
    return kg_destino(d, fechas)

def kg_grupo_dia(fecha):
    mapa = {"Mesas": "Envasadoras", "Mallas": "Mallas", "Graneleras": "Graneleras", "Industria": "Industria"}
    out = defaultdict(float)
    for p in prods([fecha]):
        g = mapa.get(p["zona"])
        if g:
            out[g] += p["kg"]
    return out

def arranque_dia(f):
    return sum(1 for p in DM[f]["detallePersonas"] if p["tipo"] == "tratamiento")

GRUPOS = ["Mallas", "Envasadoras", "Graneleras", "Industria"]

wb = Workbook()

# ═══ Por día ═════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Por día"
titulo(ws, f"Rendimiento por día — semanas {NP} (referencia) y {NA}",
       "Producción real = kg calibrador − fruta de mujeres − reciclado de mallas (misma fórmula que la app) · kg/persona sobre presentes que computan", 19)
cabecera(ws, 4, ["Fecha", "Kg calibrador", "Producción real", "Mujeres kg", "Reciclado kg", "Podrido kg", "Industria kg",
                 "Precalibrado kg", "Personas", "Computan kg/p", "Horas trabajadas", "kg / persona", "kg / hora·persona",
                 "t/h línea", "t/h máquina", "Pasadas", "Horas línea", "Palets brutos kg", "% aprovechado a palets"])

def fila_dia(r, f):
    dia = DM[f]
    kgd = kg_dest([f])
    c1 = poner(ws, r, 1, ETQ[f])
    if dia["sinParte"] and f in S_ACT:
        comentario(c1, "Día PROVISIONAL: sin parte diario todavía. Se completa al regenerar mañana.")
    poner(ws, r, 2, round(dia["kgCalibrador"]), FMT_KG)
    poner(ws, r, 3, round(dia["produccionReal"]), FMT_KG)
    poner(ws, r, 4, round(dia["mujeres"]), FMT_KG)
    poner(ws, r, 5, round(dia["z1"] + dia["z2"]), FMT_KG)
    poner(ws, r, 6, round(kgd.get("podrido", 0) + kgd.get("muestra", 0)), FMT_KG)
    poner(ws, r, 7, round(kgd.get("industria", 0)), FMT_KG)
    poner(ws, r, 8, round(kgd.get("prec", 0)), FMT_KG)
    poner(ws, r, 9, dia["presentes"])
    poner(ws, r, 10, dia["computables"])
    poner(ws, r, 11, dia["horasReales"], FMT_H)
    poner(ws, r, 12, f"=IFERROR(C{r}/J{r},\"\")", FMT_KG)
    poner(ws, r, 13, f"=IFERROR(C{r}/K{r},\"\")", FMT_KG)
    poner(ws, r, 14, dia["thEfectiva"], FMT_TH)
    poner(ws, r, 15, dia["thMaquina"], FMT_TH)
    poner(ws, r, 16, dia["nPasadas"])
    poner(ws, r, 17, dia["horasLinea"], FMT_H)
    poner(ws, r, 18, round(dia["paletsBrutos"]) if dia["paletsBrutos"] else None, FMT_KG)
    poner(ws, r, 19, f"=IFERROR(R{r}/B{r},\"\")", FMT_PCT)

def fila_total(r, r0, r1, etiqueta):
    ndias = r1 - r0 + 1
    poner(ws, r, 1, etiqueta, bold=True, fill=FILL_TOTAL)
    for c in [2, 3, 4, 5, 6, 7, 8, 11, 16, 17, 18]:
        L = get_column_letter(c)
        fmt = FMT_KG if c in (2, 3, 4, 5, 6, 7, 8, 18) else (FMT_H if c in (11, 17) else '#,##0')
        poner(ws, r, c, f"=SUM({L}{r0}:{L}{r1})", fmt, bold=True, fill=FILL_TOTAL)
    for c in (9, 10):
        L = get_column_letter(c)
        celda = poner(ws, r, c, f"=AVERAGE({L}{r0}:{L}{r1})", '0.0', bold=True, fill=FILL_TOTAL)
        comentario(celda, "Personas de MEDIA al día (no la suma de la semana).")
    poner(ws, r, 12, f"=IFERROR(C{r}/(J{r}*{ndias}),\"\")", FMT_KG, bold=True, fill=FILL_TOTAL)
    comentario(ws.cell(row=r, column=12), "kg/persona del día normal: producción real de la semana ÷ días ÷ personas medias que computan.")
    poner(ws, r, 13, f"=IFERROR(C{r}/K{r},\"\")", FMT_KG, bold=True, fill=FILL_TOTAL)
    poner(ws, r, 14, f"=IFERROR(SUMPRODUCT(N{r0}:N{r1},Q{r0}:Q{r1})/Q{r},\"\")", FMT_TH, bold=True, fill=FILL_TOTAL)
    poner(ws, r, 15, "", fill=FILL_TOTAL)
    poner(ws, r, 19, f"=IFERROR(R{r}/B{r},\"\")", FMT_PCT, bold=True, fill=FILL_TOTAL)
    for c in range(1, 20):
        ws.cell(row=r, column=c).border = BORDE_TOTAL

r = 5
poner(ws, r, 1, f"SEMANA {NP} · {rango_texto(S_ANT)} (referencia)", bold=True, fill=FILL_CAB2)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19); r += 1
rA0 = r
for f in S_ANT:
    fila_dia(r, f); r += 1
rA_t = r; fila_total(r, rA0, rA_t - 1, f"TOTAL S{NP}"); r += 2
poner(ws, r, 1, f"SEMANA {NA} · {rango_texto(S_ACT)} (informe)", bold=True, fill=FILL_CAB2)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19); r += 1
rB0 = r
for f in S_ACT:
    fila_dia(r, f); r += 1
rB_t = r; fila_total(r, rB0, rB_t - 1, f"TOTAL S{NA}")
nota(ws, rB_t + 2, "kg/persona usa la producción real y las personas que computan (fuera oficina y carga/descarga). 'kg/hora·persona' divide por las horas reales del reloj: quita el efecto de jornadas más largas o cortas.", 19)
anchos(ws, [20, 11, 12, 10, 10, 10, 11, 12, 9, 10, 11, 11, 11, 8, 9, 8, 9, 12, 11])
ws.freeze_panes = "B5"

# ═══ Resumen ═════════════════════════════════════════════════════════════════
rs = wb.create_sheet("Resumen", 0)
titulo(rs, f"Informe de rendimiento — Semana {NA} ({rango_texto(S_ACT)})",
       f"Lasarte Cítricos S.L. · datos hasta {ETQ.get(d['hasta'], d['hasta'])} · comparado con la semana {NP} ({rango_texto(S_ANT)})", 6)
cabecera(rs, 4, ["", "Concepto", f"Semana {NA}", f"Semana {NP}", "Δ", ""])
PD = "'Por día'!"
def pct_dest(fechas, claves):
    kgd = kg_dest(fechas)
    tot = sum(kgd.values())
    return sum(kgd.get(k, 0) for k in claves) / tot if tot else None
r = 5
for nombre, vA, vB, fmt, com in [
    ("Kg calibrados", f"={PD}B{rB_t}", f"={PD}B{rA_t}", FMT_KG, None),
    ("Producción real (kg)", f"={PD}C{rB_t}", f"={PD}C{rA_t}", FMT_KG, None),
    ("Personas medias/día", f"={PD}I{rB_t}", f"={PD}I{rA_t}", '0.0', None),
    ("· de ellas computan kg/persona", f"={PD}J{rB_t}", f"={PD}J{rA_t}", '0.0', None),
    ("Horas trabajadas (semana)", f"={PD}K{rB_t}", f"={PD}K{rA_t}", FMT_H, None),
    ("kg/persona (día normal)", f"={PD}L{rB_t}", f"={PD}L{rA_t}", FMT_KG, "La vara del informe semanal del lunes: producción real / personas que computan."),
    ("kg por hora·persona", f"={PD}M{rB_t}", f"={PD}M{rA_t}", FMT_KG, None),
    ("Velocidad de línea t/h (media pond.)", f"={PD}N{rB_t}", f"={PD}N{rA_t}", FMT_TH, None),
    ("% podrido + muestras", pct_dest(S_ACT, ["podrido", "muestra"]), pct_dest(S_ANT, ["podrido", "muestra"]), FMT_PCT, None),
    ("% industria", pct_dest(S_ACT, ["industria"]), pct_dest(S_ANT, ["industria"]), FMT_PCT, None),
    ("% precalibrado (vuelve a línea)", pct_dest(S_ACT, ["prec"]), pct_dest(S_ANT, ["prec"]), FMT_PCT, None),
    ("% a Mercadona (4 formatos)", pct_dest(S_ACT, ["mdna3", "mdna4", "mdna5", "mdnaGranel"]), pct_dest(S_ANT, ["mdna3", "mdna4", "mdna5", "mdnaGranel"]), FMT_PCT, None),
    ("% otros clientes", pct_dest(S_ACT, ["otrosEmp", "otrosGranel"]), pct_dest(S_ANT, ["otrosEmp", "otrosGranel"]), FMT_PCT, None),
]:
    poner(rs, r, 2, nombre)
    c = poner(rs, r, 3, vA, fmt)
    if com:
        comentario(c, com)
    poner(rs, r, 4, vB, fmt)
    poner(rs, r, 5, f"=IFERROR(C{r}/D{r}-1,\"\")", FMT_PCT)
    r += 1
r += 1
nota(rs, r, "· Antes de leer el kg/persona como productividad, mira la mezcla (% podrido, industria, precalibrado): la fruta que viene del campo manda tanto como la gente. Lee la hoja 'Cómo se mide'.", 6); r += 1
# Las gráficas van DEBAJO de la tabla (no al lado): así entran en el papel impreso.
anc1, anc2, anc3 = r + 1, r + 17, r + 33
fin_resumen = anc3 + 18

ch = BarChart(); ch.type = "col"; ch.style = 10
ch.title = "Kg calibrados por día"
cats = Reference(ws, min_col=1, min_row=rB0, max_row=rB_t - 1)
ch.add_data(Reference(ws, min_col=2, min_row=rA0, max_row=rA_t - 1), titles_from_data=False)
ch.add_data(Reference(ws, min_col=2, min_row=rB0, max_row=rB_t - 1), titles_from_data=False)
ch.set_categories(cats)
ch.series[0].tx = SeriesLabel(v=f"Semana {NP}")
ch.series[1].tx = SeriesLabel(v=f"Semana {NA}")
ch.series[0].graphicalProperties.solidFill = COLOR["gris_ref"]
ch.series[1].graphicalProperties.solidFill = COLOR["mdna"]
ch.gapWidth = 60; ch.width = 15; ch.height = 7.5
rs.add_chart(ch, f"B{anc1}")

ch2 = LineChart()
ch2.title = "kg por persona (computable) por día"
ch2.add_data(Reference(ws, min_col=12, min_row=rA0, max_row=rA_t - 1), titles_from_data=False)
ch2.add_data(Reference(ws, min_col=12, min_row=rB0, max_row=rB_t - 1), titles_from_data=False)
ch2.set_categories(cats)
ch2.series[0].tx = SeriesLabel(v=f"Semana {NP}")
ch2.series[1].tx = SeriesLabel(v=f"Semana {NA}")
ch2.series[0].graphicalProperties.line.solidFill = COLOR["gris_ref"]
ch2.series[1].graphicalProperties.line.solidFill = COLOR["mdna"]
ch2.series[0].graphicalProperties.line.width = 25000
ch2.series[1].graphicalProperties.line.width = 25000
ch2.width = 15; ch2.height = 7.5
rs.add_chart(ch2, f"B{anc2}")
anchos(rs, [2, 42, 13, 13, 9, 2])

# ═══ Cómo se mide a la gente ═════════════════════════════════════════════════
wq = wb.create_sheet("Cómo se mide", 1)
titulo(wq, "Cómo se mide a la gente y a la producción — reglas de este informe",
       "Para el jefe y la encargada: qué significa cada número, quién cuenta y qué NO depende de la gente", 10)
r = 4
def bloq(t):
    global r
    poner(wq, r, 1, t, bold=True, fill=FILL_CAB2, borde=False)
    wq.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
def lin(t):
    global r
    celda = wq.cell(row=r, column=1, value=t)
    celda.font = F_NOTA
    celda.alignment = Alignment(wrap_text=True, vertical="top")
    wq.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    wq.row_dimensions[r].height = 13 * (1 + len(t) // 130)
    r += 1

bloq("Los cuatro números que miden el día")
lin("· PRODUCCIÓN REAL = kg que pesó el calibrador − fruta apartada por las mujeres − reciclado de mallas. Es la fruta que de verdad salió adelante ese día (la misma fórmula del informe del lunes).")
lin("· KG/PERSONA = producción real ÷ personas que computan. Computan todos los presentes menos oficina, carga y descarga y el turno de tarde (lo dice la ficha de cada trabajador en la app).")
EST = cargar_estandar()
if EST and EST.get("regimenes"):
    _c = EST["regimenes"]["completa"]; _r = EST["regimenes"]["reducida"]
    lin(f"· ESTÁNDAR ACORDADO ({EST['fecha']}), POR RÉGIMEN DE PLANTILLA: media plantilla (≤{EST.get('cortePlantillaReducida', 35)} presentes) suelo {_r['kgPersonaSuelo']:,} · objetivo {_r['kgPersonaObjetivo']:,}; plantilla completa suelo {_c['kgPersonaSuelo']:,} · objetivo {_c['kgPersonaObjetivo']:,} kg/persona. Verde = en objetivo; ámbar = entre suelo y objetivo; rojo = bajo el suelo, día a explicar (¿fruta, máquina o gente?). {EST['nota']}".replace(",", "."))
elif EST:
    lin(f"· ESTÁNDAR ACORDADO ({EST['fecha']}): suelo {EST['kgPersonaSuelo']:,} · objetivo {EST['kgPersonaObjetivo']:,} kg/persona. Verde = en objetivo; ámbar = entre suelo y objetivo; rojo = bajo el suelo, día a explicar (¿fruta, máquina o gente?). {EST['nota']}".replace(",", "."))
lin("· KG/HORA·PERSONA = producción real ÷ horas reales fichadas. Es la medida justa cuando las jornadas son desiguales: un viernes de 8:45 h no puede compararse en kg/persona con un día de 7 h.")
lin("· T/H DE LÍNEA = kg ÷ horas de pasada del calibrador. Mide la máquina y su alimentación, no a las personas.")
bloq("La regla de las zonas (acordada el 14-08-2026)")
lin("· Cada zona (mallas, envasadoras, graneleras) cuenta SUS personas MÁS las del ARRANQUE: encargadas, producción, aéreo, tría de podrido, volcador, carretillas, transpaletas y mantenimiento. Sin arranque la línea no corre, así que su gente cuenta en todas las zonas.")
lin("· La INDUSTRIA la produce el propio arranque (la fruta cae sola al box): sus kg/persona se calculan solo con el arranque. Nunca aparece con 0 personas.")
lin("· Por eso las zonas NO se pueden sumar entre sí: el arranque está en todas. Es a propósito — cada zona responde '¿cuánta gente hace falta de verdad para producir este kg?'.")
lin("· Las personas salen en MEDIAS AL DÍA (9,6 de arranque = un día faltó una). Los enteros exactos de cada día están en las tablas diarias, y los nombres en 'Quién es quién'.")
bloq("Lo que NO depende de la gente (mirar antes de juzgar un número)")
lin("· La MEZCLA de fruta manda: un día con mucho podrido, industria o verde rinde menos kg/persona sin que nadie trabaje peor. Por eso el % podrido, % industria y % precalibrado van al lado de cada kg/persona.")
lin("· El PRECALIBRADO apartado hoy no cuenta como confección de nadie hoy: contará el día que vuelva a línea y se confeccione.")
lin("· Las AVERÍAS y esperas bajan la t/h de línea, no el rendimiento de las personas. Si t/h y kg/persona caen a la vez, mirar primero la máquina y la fruta.")
bloq("De dónde sale cada dato (y cómo corregirlo)")
lin("· Kg, categorías y calibres: del propio calibrador (BD del Sizer; si un día no está volcado, de los informes DOCX que el Sizer envía al cerrar cada lote, validados contra sus totales).")
lin("· Personas y horas: del reloj de presencia (export de asistencias). Quién computa y la zona de cada uno: de la ficha del trabajador en la app — si alguien está en la zona equivocada, se corrige su ficha y el informe sale bien al regenerar.")
lin("· Día de HOY: provisional hasta mañana (el parte diario y los palets del ERP se completan por la mañana).")
bloq("Cómo se regenera (cada mañana, un clic)")
lin("· Doble clic en scripts\\informe-produccion\\generar-informes.cmd (en la carpeta del repositorio). Deja antes el export del reloj como asistencias.xlsx en esa carpeta si hay días nuevos de personas.")
lin("· Salen los dos Excel en scripts\\informe-produccion\\salida\\, siempre la semana en curso comparada con la anterior completa.")
anchos(wq, [130])

# ═══ Destino de la naranja ═══════════════════════════════════════════════════
wd = wb.create_sheet("Destino de la naranja")
titulo(wd, f"Dónde fue cada kg — semana {NA} por día (contraste: total semana {NP})", "Kg del calibrador clasificados por destino comercial (misma regla que Económico → Rentabilidad)", 6 + len(S_ACT))
cabs = ["Destino"] + [ETQ[f] for f in S_ACT] + [f"Total S{NA}", f"% S{NA}", f"Total S{NP}", f"% S{NP}", "Δ p.p."]
cabecera(wd, 4, cabs)
destA, _ = destino_kg_dia(d, set(S_ACT))
destP, _ = destino_kg_dia(d, set(S_ANT))
orden = d["constantes"]["destinosOrden"]
nD = len(S_ACT)
cT, cPct, cT2, cPct2, cDelta = 2 + nD, 3 + nD, 4 + nD, 5 + nD, 6 + nD
r = 5; r0 = r
rT = r0 + len(orden)
for dst in orden:
    poner(wd, r, 1, DESTINO_ES[dst])
    for j, f in enumerate(S_ACT):
        poner(wd, r, 2 + j, round(destA.get(dst, {}).get(f, 0)), FMT_KG)
    L0, L1 = get_column_letter(2), get_column_letter(1 + nD)
    poner(wd, r, cT, f"=SUM({L0}{r}:{L1}{r})", FMT_KG)
    poner(wd, r, cPct, f"={get_column_letter(cT)}{r}/{get_column_letter(cT)}${rT}", FMT_PCT)
    poner(wd, r, cT2, round(sum(destP.get(dst, {}).values())), FMT_KG)
    poner(wd, r, cPct2, f"={get_column_letter(cT2)}{r}/{get_column_letter(cT2)}${rT}", FMT_PCT)
    poner(wd, r, cDelta, f"={get_column_letter(cPct)}{r}-{get_column_letter(cPct2)}{r}", '0.0%;[Red]-0.0%')
    r += 1
poner(wd, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
for cidx in list(range(2, 2 + nD)) + [cT, cT2]:
    L = get_column_letter(cidx)
    poner(wd, rT, cidx, f"=SUM({L}{r0}:{L}{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
poner(wd, rT, cPct, 1, FMT_PCT, fill=FILL_TOTAL); poner(wd, rT, cPct2, 1, FMT_PCT, fill=FILL_TOTAL)
anchos(wd, [26] + [10] * nD + [11, 8, 11, 8, 8])

# gráfica apilada de destinos (datos a la derecha del Resumen)
gr = wb["Resumen"]
ch3 = BarChart(); ch3.type = "col"; ch3.grouping = "stacked"; ch3.overlap = 100
ch3.title = f"Semana {NA} — destino por día (kg)"
grupos_chart = [
    ("Mercadona", ["mdna3", "mdna4", "mdna5", "mdnaGranel"], COLOR["mdna"]),
    ("Otros clientes", ["otrosEmp", "otrosGranel"], COLOR["otros"]),
    ("Precalibrado", ["prec"], COLOR["prec"]),
    ("Industria", ["industria"], COLOR["industria"]),
    ("Podrido+muestras", ["podrido", "muestra"], COLOR["podrido"]),
]
base = 40
COL_LBL = 14
anc3_res = d.get("_anc3", 55)
gr.cell(row=base - 1, column=COL_LBL, value="(datos de la gráfica de destinos)").font = F_NOTA
for j, f in enumerate(S_ACT):
    gr.cell(row=base, column=COL_LBL + 1 + j, value=ETQ[f]).font = F_NOTA
for i, (nombre, claves, color) in enumerate(grupos_chart):
    gr.cell(row=base + 1 + i, column=COL_LBL, value=nombre).font = F_NOTA
    for j, f in enumerate(S_ACT):
        cell = gr.cell(row=base + 1 + i, column=COL_LBL + 1 + j, value=round(sum(destA.get(k, {}).get(f, 0) for k in claves)))
        cell.font = F_NOTA
        cell.number_format = FMT_KG
datac = Reference(gr, min_col=COL_LBL, min_row=base + 1, max_col=COL_LBL + nD, max_row=base + len(grupos_chart))
catsc = Reference(gr, min_col=COL_LBL + 1, min_row=base, max_col=COL_LBL + nD, max_row=base)
ch3.add_data(datac, titles_from_data=True, from_rows=True)
ch3.set_categories(catsc)
for i, (_, _, color) in enumerate(grupos_chart):
    ch3.series[i].graphicalProperties.solidFill = color
ch3.width = 15; ch3.height = 8
gr.add_chart(ch3, f"B{anc3_res}")

# ═══ Categorías y calidad ════════════════════════════════════════════════════
def hoja_matriz(nombre_hoja, titulo_txt, sub, filas_dict, nombre_col):
    wc = wb.create_sheet(nombre_hoja)
    titulo(wc, titulo_txt, sub, 6 + len(S_ACT))
    cabecera(wc, 4, [nombre_col] + [ETQ[f] for f in S_ACT] + [f"Total S{NA}", f"% S{NA}", f"Total S{NP}", f"% S{NP}", "Δ p.p."])
    claves = sorted(filas_dict.keys(), key=lambda k: -sum(v for f, v in filas_dict[k].items() if f in S_ACT))
    n = len(claves)
    r = 5; r0 = r; rT = r0 + n
    for cl in claves:
        poner(wc, r, 1, cl)
        for j, f in enumerate(S_ACT):
            poner(wc, r, 2 + j, round(filas_dict[cl].get(f, 0)), FMT_KG)
        L0, L1 = get_column_letter(2), get_column_letter(1 + nD)
        poner(wc, r, cT, f"=SUM({L0}{r}:{L1}{r})", FMT_KG)
        poner(wc, r, cPct, f"=IFERROR({get_column_letter(cT)}{r}/{get_column_letter(cT)}${rT},\"\")", FMT_PCT)
        poner(wc, r, cT2, round(sum(v for f, v in filas_dict[cl].items() if f in S_ANT)), FMT_KG)
        poner(wc, r, cPct2, f"=IFERROR({get_column_letter(cT2)}{r}/{get_column_letter(cT2)}${rT},\"\")", FMT_PCT)
        poner(wc, r, cDelta, f"=IFERROR({get_column_letter(cPct)}{r}-{get_column_letter(cPct2)}{r},\"\")", '0.0%;[Red]-0.0%')
        r += 1
    poner(wc, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
    for cidx in list(range(2, 2 + nD)) + [cT, cT2]:
        L = get_column_letter(cidx)
        poner(wc, rT, cidx, f"=SUM({L}{r0}:{L}{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
    anchos(wc, [16] + [10] * nD + [11, 8, 11, 8, 8])
    return wc, rT

cat = defaultdict(lambda: defaultdict(float))
for c in d["categoriasDia"]:
    cat[c["clase"]][c["fecha"]] += c["kg"]
wc, rTc = hoja_matriz("Categorías y calidad", f"Reparto por categoría/calidad del calibrador — semana {NA}",
                      f"Kg pesados por el calibrador en cada clase · contraste con la semana {NP}", cat, "Clase")
nota(wc, rTc + 2, "'Mujeres' es la fruta que se aparta en el aéreo y vuelve a pasar otro día; 'Densidad' es descarte automático a industria. Extra/Cat1/Cat 2 son calidades de venta; Cat 3 y Verde alimentan precalibrado e industria.", 6 + nD)

cal = defaultdict(lambda: defaultdict(float))
for c in d["calibresDia"]:
    cal[c["tamano"]][c["fecha"]] += c["kg"]
hoja_matriz("Calibres", f"Reparto por calibre (tamaño) — semana {NA}",
            f"Kg por tamaño del calibrador · contraste con la semana {NP}", cal, "Calibre")

# ═══ Zonas: kg por persona (regla del dueño 14-08) ═══════════════════════════
wz = wb.create_sheet("Zonas y kg-persona")
titulo(wz, f"Kg por persona en cada zona — semana {NA} (personas de la zona + arranque)",
       "El arranque (encargadas, producción, aéreo, tría podrido, volcador, carretillas, transpaletas, mantenimiento) cuenta en TODAS las zonas: sin él la línea no corre. Industria = solo el arranque.", 9)
cabecera(wz, 4, ["Fecha", "Zona", "Kg de la zona", "Personas de la zona", "+ Arranque", "Total personas", "kg / persona", "", ""])
r = 5
for f in S_ACT:
    dia = DM[f]
    kgg = kg_grupo_dia(f)
    arr = arranque_dia(f)
    for g in GRUPOS:
        kgz = kgg.get(g, 0)
        pz = dia["personasPorGrupo"].get(g, 0)
        if kgz <= 0 and pz <= 0:
            continue
        poner(wz, r, 1, ETQ[f])
        poner(wz, r, 2, g if g != "Industria" else "Industria (arranque)")
        poner(wz, r, 3, round(kgz), FMT_KG)
        poner(wz, r, 4, pz)
        poner(wz, r, 5, arr)
        poner(wz, r, 6, f"=D{r}+E{r}")
        poner(wz, r, 7, f"=IFERROR(C{r}/F{r},\"—\")", FMT_KG)
        r += 1
r += 1
wz.row_breaks.append(Break(id=r - 1))  # al imprimir: pág. 1 = detalle diario, pág. 2 = resumen + quién es quién
poner(wz, r, 1, f"RESUMEN SEMANA {NA} POR ZONA — un día normal de esta semana", bold=True, fill=FILL_CAB2)
wz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
cabecera(wz, r, ["Zona", "Kg al día", "Personas de la zona", "+ Arranque", "Total personas", "kg / persona", f"S{NP}: kg/persona", "Δ", "Kg semana entera"], oscura=False)
r += 1
nA_d = len(S_ACT); nP_d = max(len(S_ANT), 1)
arrAm = sum(arranque_dia(f) for f in S_ACT) / nA_d
arrPm = sum(arranque_dia(f) for f in S_ANT) / nP_d if S_ANT else 0
for g in GRUPOS:
    kgA = sum(kg_grupo_dia(f).get(g, 0) for f in S_ACT)
    pzAm = sum(DM[f]["personasPorGrupo"].get(g, 0) for f in S_ACT) / nA_d
    kgP = sum(kg_grupo_dia(f).get(g, 0) for f in S_ANT)
    pzPm = sum(DM[f]["personasPorGrupo"].get(g, 0) for f in S_ANT) / nP_d if S_ANT else 0
    kgpP = (kgP / nP_d) / (pzPm + arrPm) if (pzPm + arrPm) > 0 else None
    poner(wz, r, 1, g if g != "Industria" else "Industria (arranque)", bold=True)
    poner(wz, r, 2, round(kgA / nA_d), FMT_KG)
    poner(wz, r, 3, round(pzAm, 1), '0.0')
    poner(wz, r, 4, round(arrAm, 1), '0.0')
    poner(wz, r, 5, f"=C{r}+D{r}", '0.0')
    poner(wz, r, 6, f"=IFERROR(B{r}/E{r},\"—\")", FMT_KG)
    poner(wz, r, 7, kgpP, FMT_KG)
    poner(wz, r, 8, f"=IFERROR(F{r}/G{r}-1,\"\")", FMT_PCT)
    poner(wz, r, 9, round(kgA), FMT_KG)
    r += 1
nota(wz, r + 1, "Todo en 'un día normal': kg al día = kg de la semana ÷ días, y las personas son las de media al día (un 9,6 = un día faltó una). El detalle diario exacto está arriba con números enteros.", 9)
nota(wz, r + 2, "Las zonas no se suman entre sí: el arranque cuenta en todas a propósito — sin él la línea no corre. La fruta de industria la produce el propio arranque.", 9)
r += 4
poner(wz, r, 1, "QUIÉN ES QUIÉN ESTA SEMANA (días presentes entre paréntesis)", bold=True, fill=FILL_CAB2)
wz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
quien = {}
for f in S_ACT:
    for p in DM[f]["detallePersonas"]:
        clave = p["grupo"] or ("Arranque (línea)" if p["tipo"] == "tratamiento" else ("No computan (carga y descarga / tarde)" if p["tipo"] == "no_computa" else "Otros"))
        quien.setdefault(clave, {}).setdefault(p["nombre"], 0)
        quien[clave][p["nombre"]] += 1
for grupo in ["Arranque (línea)", "Mallas", "Envasadoras", "Graneleras", "Industria", "No computan (carga y descarga / tarde)", "Otros"]:
    personas = quien.get(grupo)
    if not personas:
        continue
    poner(wz, r, 1, f"{grupo} — {len(personas)}", bold=True)
    lista = "; ".join(f"{n} ({v})" for n, v in sorted(personas.items()))
    c = poner(wz, r, 2, lista)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wz.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    wz.row_dimensions[r].height = 15 * (1 + len(lista) // 95)
    r += 1
anchos(wz, [24, 16, 12, 12, 12, 13, 13, 10, 12])

# ═══ Salidas por cliente (ERP) — SIN dinero: el € va solo al informe económico ═
wv = wb.create_sheet("Salidas por cliente")
titulo(wv, f"Palets por cliente y día (alta en el ERP) — semana {NA}", "Kilos salidos de línea por cliente · el día de HOY va hasta la última sincronización del espejo (se completa mañana)", 3 + len(S_ACT))
cabecera(wv, 4, ["Cliente", "Tipo"] + [ETQ[f] for f in S_ACT] + ["Total kg"])
cli = defaultdict(lambda: {"tipo": "comercial", "dias": defaultdict(float)})
for v in d["ventasErp"]:
    if v["fecha"] not in S_ACT:
        continue
    tipo = "interno" if es_interno(v["articulo"]) else "comercial"
    nombre = v["cliente"] or ("(movimiento interno)" if tipo == "interno" else "(sin cliente aún)")
    a = cli[nombre]
    a["tipo"] = tipo
    a["dias"][v["fecha"]] += v["kg"]
r = 5; r0 = r
cTot = 3 + nD
for nombre, a in sorted(cli.items(), key=lambda x: -sum(x[1]["dias"].values())):
    poner(wv, r, 1, nombre)
    poner(wv, r, 2, a["tipo"])
    for j, f in enumerate(S_ACT):
        poner(wv, r, 3 + j, round(a["dias"].get(f, 0)) or None, FMT_KG)
    L0, L1 = get_column_letter(3), get_column_letter(2 + nD)
    poner(wv, r, cTot, f"=SUM({L0}{r}:{L1}{r})", FMT_KG)
    r += 1
rT = r
poner(wv, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
for cidx in range(3, 4 + nD):
    L = get_column_letter(cidx)
    poner(wv, rT, cidx, f"=SUM({L}{r0}:{L}{rT-1})", FMT_KG, bold=True, fill=FILL_TOTAL)
nota(wv, rT + 2, "'(sin cliente aún)' = palets todavía sin albarán de venta: en cámara o pendientes de asignar. Los internos (PREC/CITRICA) no son venta, son movimientos del almacén.", 3 + nD)
anchos(wv, [34, 10] + [10] * nD + [11])

# ═══ Plantilla completa: comparativa ═════════════════════════════════════════
wp = wb.create_sheet("Plantilla completa")
mejores = sorted(TODOS, key=lambda f: (-DM[f]["presentes"], f))[:2]
maxp = max(DM[f]["presentes"] for f in mejores) if mejores else 0
titulo(wp, f"Comparativa con plantilla al completo ({' y '.join(ETQ[f] for f in mejores)}, {maxp} personas)",
       f"Los dos días con más gente de las dos semanas, comparados con cada día de la semana {NA}", 12)
cabecera(wp, 4, ["", "Kg calibrador", "Producción real", "Personas", "Computan", "Horas", "kg/persona", "kg/h·persona", "t/h línea", "% podrido", "% industria", "% Mercadona"])
r = 5
def fila_comp(r, f, etiqueta, fill=None):
    dia = DM[f]
    kgd = kg_dest([f]); tot = sum(kgd.values()) or 1
    poner(wp, r, 1, etiqueta, bold=fill is not None, fill=fill)
    poner(wp, r, 2, round(dia["kgCalibrador"]), FMT_KG, fill=fill)
    poner(wp, r, 3, round(dia["produccionReal"]), FMT_KG, fill=fill)
    poner(wp, r, 4, dia["presentes"], None, fill=fill)
    poner(wp, r, 5, dia["computables"], None, fill=fill)
    poner(wp, r, 6, dia["horasReales"], FMT_H, fill=fill)
    poner(wp, r, 7, f"=IFERROR(C{r}/E{r},\"\")", FMT_KG, fill=fill)
    poner(wp, r, 8, f"=IFERROR(C{r}/F{r},\"\")", FMT_KG, fill=fill)
    poner(wp, r, 9, dia["thEfectiva"], FMT_TH, fill=fill)
    poner(wp, r, 10, (kgd.get("podrido", 0) + kgd.get("muestra", 0)) / tot, FMT_PCT, fill=fill)
    poner(wp, r, 11, kgd.get("industria", 0) / tot, FMT_PCT, fill=fill)
    poner(wp, r, 12, sum(kgd.get(k, 0) for k in ("mdna3", "mdna4", "mdna5", "mdnaGranel")) / tot, FMT_PCT, fill=fill)

for f in mejores:
    fila_comp(r, f, f"REFERENCIA · {ETQ[f]} (S{NP if f in S_ANT else NA})", FILL_CAB2); r += 1
rref0, rref1 = 5, r - 1
poner(wp, r, 1, "Media de referencia", bold=True, fill=FILL_TOTAL)
for cidx in range(2, 13):
    L = get_column_letter(cidx)
    fmt = FMT_KG if cidx in (2, 3, 7, 8) else (FMT_H if cidx == 6 else (FMT_TH if cidx == 9 else (FMT_PCT if cidx >= 10 else '0.0')))
    poner(wp, r, cidx, f"=AVERAGE({L}{rref0}:{L}{rref1})", fmt, bold=True, fill=FILL_TOTAL)
rMedia = r
r += 2
for f in S_ACT:
    fila_comp(r, f, ETQ[f] + f" (S{NA})"); r += 1
rS0, rS1 = rMedia + 2, r - 1
poner(wp, r, 1, f"Δ semana {NA} vs referencia", bold=True, fill=FILL_TOTAL)
for cidx in [2, 3, 6, 7, 8, 9]:
    L = get_column_letter(cidx)
    poner(wp, r, cidx, f"=IFERROR(AVERAGE({L}{rS0}:{L}{rS1})/{L}{rMedia}-1,\"\")", FMT_PCT, bold=True, fill=FILL_TOTAL)
r += 2
nota(wp, r, "Los días de referencia también tuvieron su mezcla de fruta: mira % podrido/industria antes de leer el kg/persona como pura productividad.", 12); r += 1
cabecera(wp, r + 1, ["Personas por día", "Arranque", "Mallas", "Envasadoras", "Graneleras", "No computan", "Total computan", "", "", "", "", ""], oscura=False)
r2 = r + 2
for f in mejores + S_ACT:
    dia = DM[f]
    poner(wp, r2, 1, ("REF · " if f in mejores else "") + ETQ[f])
    poner(wp, r2, 2, arranque_dia(f))
    for j, g in enumerate(["Mallas", "Envasadoras", "Graneleras"]):
        poner(wp, r2, 3 + j, dia["personasPorGrupo"].get(g, 0))
    poner(wp, r2, 6, dia["presentes"] - dia["computables"])
    poner(wp, r2, 7, dia["computables"], None, bold=True)
    r2 += 1
nota(wp, r2 + 1, "El arranque es quien corre la línea y produce la industria; cuenta dentro de 'Total computan'.", 12)
anchos(wp, [26, 11, 12, 9, 9, 8, 11, 12, 8, 9, 10, 12])

# ═══ Trabajadores ════════════════════════════════════════════════════════════
wt = wb.create_sheet("Trabajadores")
titulo(wt, f"Presencia y horas por trabajador — semana {NA}", "Horas del reloj de presencia · zona y si computa para el kg/persona (ficha de la app)", 6 + len(S_ACT))
cabecera(wt, 4, ["Trabajador", "Zona", "Grupo confección", "Computa kg/p"] + [ETQ[f] for f in S_ACT] + ["Días", "Horas semana"])
pers = {}
for f in S_ACT:
    for p in DM[f]["detallePersonas"]:
        a = pers.setdefault(p["nombre"], {"zona": p["zona"], "grupo": p["grupo"], "computa": p["computa"], "dias": {}, "conFicha": p["conFicha"], "tipo": p["tipo"]})
        a["dias"][f] = p["horas"]
r = 5; r0 = r
cDias, cHoras = 5 + nD, 6 + nD
for nombre, a in sorted(pers.items(), key=lambda x: ((x[1]["grupo"] or ("0arranque" if x[1]["tipo"] == "tratamiento" else "zz")), x[0])):
    poner(wt, r, 1, nombre + ("" if a["conFicha"] else " *"))
    poner(wt, r, 2, a["zona"])
    poner(wt, r, 3, a["grupo"] or ("Arranque" if a["tipo"] == "tratamiento" else "—"))
    poner(wt, r, 4, "Sí" if a["computa"] else "No")
    for j, f in enumerate(S_ACT):
        poner(wt, r, 5 + j, a["dias"].get(f), FMT_H)
    L0, L1 = get_column_letter(5), get_column_letter(4 + nD)
    poner(wt, r, cDias, f"=COUNT({L0}{r}:{L1}{r})")
    poner(wt, r, cHoras, f"=SUM({L0}{r}:{L1}{r})", FMT_H)
    r += 1
rT = r
poner(wt, rT, 1, "TOTAL", bold=True, fill=FILL_TOTAL)
for cidx in range(5, 7 + nD):
    L = get_column_letter(cidx)
    poner(wt, rT, cidx, f"=SUM({L}{r0}:{L}{rT-1})", FMT_H if cidx != cDias else '#,##0', bold=True, fill=FILL_TOTAL)
nota(wt, rT + 2, "* sin ficha en la app: sin zona asignada (crear su ficha de trabajador para que cuente en su zona). El turno de tarde computa horas pero no kg/persona, como en la app.", 6 + nD)
anchos(wt, [34, 24, 15, 11] + [8] * nD + [7, 11])
wt.freeze_panes = "B5"

# ═══ Metodología ═════════════════════════════════════════════════════════════
wm = wb.create_sheet("Metodología")
titulo(wm, "Definiciones y fuentes", "", 8)
r = 4
for t in [
    "· Producción real = kg del calibrador − fruta apartada por las mujeres − reciclado de mallas Z1/Z2 (fórmula de la app). El día de HOY aún no tiene parte: mujeres según el propio calibrador y reciclado desconocido.",
    "· kg/persona = producción real ÷ presentes que computan (se excluyen oficina, carga/descarga y quien tenga 'no computa' en su ficha).",
    "· kg/hora·persona = producción real ÷ horas reales fichadas.",
    "· t/h línea = kg ÷ horas de pasada (inicio→fin del lote). t/h máquina = 'Toneladas/Hora' del Sizer por pasada (media ponderada).",
    "· Categorías/calibres: pesos del calibrador por clase y tamaño (BD espejo; si un día no está volcado, DOCX del receptor validados contra sus totales).",
    "· Zonas: los kg salen de clasificar el PRODUCTO con el clasificador de la app; las personas de cada zona son las de su ficha MÁS las del arranque, y la industria la produce el arranque (regla del dueño, 14-08-2026).",
    "· Salidas por cliente: espejo del ERP (solo SELECT). Un palet 'sin cliente aún' está en cámara.",
    "· Asistencia: export del reloj (asistencias.xlsx en esta carpeta). La semana en curso no está en la app hasta el volcado del lunes.",
]:
    nota(wm, r, t, 8); r += 1
r += 1
poner(wm, r, 1, "Avisos de este periodo", bold=True, fill=FILL_CAB2)
wm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 1
# Este informe es de rendimiento (para la encargada): los avisos de dinero
# (tarifas, costes) van solo en el informe económico del jefe.
for a in d["avisos"]:
    if any(x in a for x in ("€", "facturar", "tarifa", "coste")):
        continue
    nota(wm, r, "· " + a, 8); r += 1
anchos(wm, [120])

# Maquetación de impresión: este informe se imprime y se discute en persona.
imprimir_todo(wb, sin_titulos=("Resumen", "Cómo se mide", "Metodología"), areas={"Resumen": f"A1:H{fin_resumen}"})

os.makedirs("salida", exist_ok=True)
NOMBRE = f"salida/Informe rendimiento - semana {NA} ({rango_texto(S_ACT).replace('–','-')}).xlsx"
wb.save(NOMBRE)
print("OK informe B ->", NOMBRE)
