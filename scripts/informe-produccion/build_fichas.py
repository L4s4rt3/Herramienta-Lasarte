# build_fichas.py — LA PLANTILLA EN UNA PÁGINA (decisión del dueño 17-08):
# una sola hoja con los datos objetivos de todas las personas, sin rúbrica,
# con el efecto presencia dicho en palabras llanas. La valoración se habla.
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from comun_informes import MESES, preparar_impresion, anchos

with open("salida/fichas-personas.json", encoding="utf-8") as f:
    d = json.load(f)

MES = d["mes"]
anio, mes_n = MES.split("-")
MES_TXT = f"{MESES[int(mes_n) - 1]} de {anio}"

# ─── estilo propio de esta hoja (más limpio que las tablas de informe) ────────
TINTA = "1A1A2E"
GRIS = "6B7280"
F_T1 = Font(name="Arial", size=16, bold=True, color=TINTA)
F_T2 = Font(name="Arial", size=9, color=GRIS)
F_GRUPO = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_CABZ = Font(name="Arial", size=9, bold=True, color=TINTA)
F_P = Font(name="Arial", size=10, color=TINTA)
F_P_B = Font(name="Arial", size=10, bold=True, color=TINTA)
F_MAL = Font(name="Arial", size=10, bold=True, color="93384F")
F_SUAVE = Font(name="Arial", size=10, color=GRIS)
FILL_GRUPO = PatternFill("solid", fgColor="35415C")
FILL_ZEBRA = PatternFill("solid", fgColor="F4F6F9")
LINEA = Border(bottom=Side(style="hair", color="D9DDE3"))
LINEA_CAB = Border(bottom=Side(style="medium", color=TINTA))

def hhmm(minutos):
    if minutos is None:
        return "—"
    return f"{int(minutos // 60):02d}:{int(minutos % 60):02d}"

def se_nota(f):
    """El efecto presencia en palabras llanas, sin jerga."""
    e = f["efecto"]
    if not f["computa"]:
        return ("—", F_SUAVE)
    if not e["medible"]:
        return (("casi no falta", F_SUAVE) if e["nAusente"] <= 1 else ("pocas faltas para saberlo", F_SUAVE))
    delta = (e["idxAusente"] - e["idxPresente"]) * 100
    if e["nAusente"] >= 5 and delta <= -10:
        return ("SÍ — baja mucho cuando falta", F_P_B)
    if e["nAusente"] >= 5 and delta <= -5:
        return ("sí — se nota cuando falta", F_P_B)
    if e["nAusente"] >= 5 and delta >= 5:
        return ("al revés — preguntar", F_SUAVE)
    return ("no se aprecia diferencia", F_P)

fichas = [f for f in d["fichas"] if f["mes"]["diasPresente"] > 0 or f["mes"]["faltas"] > 0]

# horario propio: quien entra sistemáticamente a otra hora que el grueso del
# equipo (mozos, turno de tarde) no puede "llegar tarde" contra la mediana
entradas_equipo = sorted(f["mes"]["entradaMedia"] for f in fichas if f["mes"]["entradaMedia"] is not None)
mediana_equipo = entradas_equipo[len(entradas_equipo) // 2] if entradas_equipo else None
def horario_propio(f):
    e = f["mes"]["entradaMedia"]
    return mediana_equipo is not None and e is not None and abs(e - mediana_equipo) > 45
orden_grupo = ["Arranque", "Mallas", "Envasadoras", "Graneleras", "Industria", None]
etiqueta_grupo = {"Arranque": "ARRANQUE (los que corren la línea)", "Mallas": "MALLAS", "Envasadoras": "ENVASADORAS Y MOZOS", "Graneleras": "GRANELERAS", "Industria": "INDUSTRIA", None: "RESTO (carga y descarga, tarde, sin zona)"}

wb = Workbook()
ws = wb.active
ws.title = "Plantilla"

ws["A1"] = f"La plantilla — {MES_TXT}"
ws["A1"].font = F_T1
ws.merge_cells("A1:I1")
ws["A2"] = f"Datos automáticos del reloj y de la producción · {d['diasProduccionMes']} días de producción en el mes · sin euros · la valoración se habla en persona, esto es la base"
ws["A2"].font = F_T2
ws.merge_cells("A2:I2")
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 14

CABS = ["", "Días", "Faltas", "Extra", "Horas/día", "Entrada", "Tarde", "¿Se nota cuando falta?", ""]
r = 4
for j, c in enumerate(CABS, start=1):
    celda = ws.cell(row=r, column=j, value=c)
    celda.font = F_CABZ
    celda.border = LINEA_CAB
    celda.alignment = Alignment(horizontal="left" if j in (1, 8) else "center", vertical="center")
r += 1

zebra = False
for grupo in orden_grupo:
    del_grupo = [f for f in fichas if f["grupo"] == grupo]
    if not del_grupo:
        continue
    banda = ws.cell(row=r, column=1, value=etiqueta_grupo[grupo])
    banda.font = F_GRUPO
    banda.fill = FILL_GRUPO
    banda.alignment = Alignment(vertical="center")
    for j in range(2, 10):
        ws.cell(row=r, column=j).fill = FILL_GRUPO
    ws.row_dimensions[r].height = 16
    r += 1
    zebra = False
    for f in sorted(del_grupo, key=lambda x: x["nombre"]):
        m = f["mes"]
        texto, fuente = se_nota(f)
        fila = [
            f["nombre"],
            f"{m['diasPresente']} de {m['diasPosibles']}",
            m["faltas"] or "",
            m["sabados"] or "",
            f"{m['mediaHoras']:.1f}" if m["mediaHoras"] is not None else "—",
            hhmm(m["entradaMedia"]),
            "propio" if horario_propio(f) else (m["tarde"] if m["tardeN"] else "—"),
            texto,
            "",
        ]
        for j, v in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=j, value=v)
            celda.font = F_P
            celda.border = LINEA
            celda.alignment = Alignment(horizontal="left" if j in (1, 8) else "center", vertical="center")
            if zebra:
                celda.fill = FILL_ZEBRA
        ws.cell(row=r, column=1).font = F_P_B
        ws.cell(row=r, column=8).font = fuente
        if (m["faltas"] or 0) >= 3:
            ws.cell(row=r, column=3).font = F_MAL
        if m["tardeN"] and m["tarde"] >= 3 and not horario_propio(f):
            ws.cell(row=r, column=7).font = F_MAL
        if horario_propio(f):
            ws.cell(row=r, column=7).font = F_SUAVE
        ws.row_dimensions[r].height = 15
        zebra = not zebra
        r += 1

r += 1
for txt in [
    "Cómo leerlo — Días: trabajados de los posibles del mes en su periodo. Extra: sábados u otros días fuera de producción. Entrada: hora media. Tarde: días entrando >10 min después que el resto del equipo («propio» = su turno empieza a otra hora: mozos, tarde).",
    "«¿Se nota cuando falta?»: compara cómo rinde el almacén entero (kg por persona) los días que esa persona viene y los días que falta, desde mayo. Si pone «sí», los días sin ella/él el almacén rinde claramente menos.",
    "«Al revés — preguntar» no es una acusación: con pocas faltas puede ser casualidad (la fruta de esos días). «Casi no falta» es la mejor señal que hay.",
]:
    celda = ws.cell(row=r, column=1, value=txt)
    celda.font = F_T2
    celda.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 11 * (1 + len(txt) // 125)
    r += 1

anchos(ws, [34, 9, 7, 7, 9, 9, 7, 30, 2])
ws.sheet_view.showGridLines = False
preparar_impresion(ws, titulo_filas=None, una_pagina=True, vertical=True)

os.makedirs("salida", exist_ok=True)
NOMBRE = f"salida/PLANTILLA - {MES}.xlsx"
wb.save(NOMBRE)
print(f"OK plantilla en una página ({len(fichas)} personas) -> {NOMBRE}")
